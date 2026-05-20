"""
src/annex_iv.py
===============
Annex IV regulatory transparency report — AIFMD Article 110 / EU 231/2013.
Quarterly submission to the CSSF for all AIFM funds.

Usage
-----
    from src.annex_iv import build_annex_iv, export_annex_iv_excel

    rpt  = build_annex_iv(engine, 'AIFM_HedgeFund', quarter='2026-03-31')
    rpt['identification']       # fund ID fields — DataFrame(field, value)
    rpt['asset_class_breakdown']
    rpt['risk_measures']
    rpt['leverage_detail']
    rpt['liquidity_buckets']

    path = export_annex_iv_excel(engine, quarter='2026-03-31')

Output (Excel)
--------------
    data/annex_iv_report_<quarter>.xlsx

    Sheets: Summary | AIFM_HedgeFund | AIFM_PrivateDebt | AIFM_RealEstate
            | AIFM_PE_Buyout | AIFM_Infra_Core

Regulatory basis
----------------
    AIFMD Article 110 — Annex IV transparency reporting
    EU 231/2013 Articles 110-121 — AIFM regulatory reporting requirements
    ESMA technical guidance v1.7 (July 2024) — Annex IV field definitions
    AIFMD II (Directive 2024/927/EU) — expanded LMT and delegation disclosures
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Any, Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from src.database import (
    FUND_METADATA,
    InfraAsset,
    InfraDebt,
    InfraFund,
    InfraFundInvestment,
    InfraNavHistory,
    PECashFlow,
    PEFund,
    PEFundCashManagement,
    PEFundInvestment,
    PENavHistory,
    PEPortfolioCompany,
    get_engine,
    query_nav_history,
)
from src.enrichment import get_risk_ready_df
from src.infra_utils import (
    asset_nav_breakdown,
    concentration_by_sector,
    duration_profile,
    fund_nav_timeseries,
    inflation_sensitivity,
    infra_irr,
    infra_multiples,
)
from src.leverage_config import INSTRUMENT_SOURCE
from src.pe_utils import fund_irr, pe_multiples
from src.risk_utils import (
    days_to_liquidate,
    es_historical,
    investor_concentration,
    liquidity_buckets,
    var_historical,
    var_scale,
)

# ── static configuration ──────────────────────────────────────────────────

_VALUATION_DATE = '2026-05-13'   # static snapshot date for liquid funds
_AIFM_NAME      = 'ManCo SA'
_AIFM_LEI       = '213800SIMULATED0001'
_DEPOSITARY     = 'BNP Paribas Securities Services Luxembourg'
_ADMINISTRATOR  = 'IQ-EQ Luxembourg'
_AUDITOR        = 'PricewaterhouseCoopers Luxembourg'

_STRATEGY = {
    'AIFM_HedgeFund' : 'Long/Short Equity',
    'AIFM_PrivateDebt': 'Private Debt — Direct Lending',
    'AIFM_RealEstate' : 'Core Real Estate',
    'AIFM_PE_Buyout'   : 'Leveraged Buyout / Growth Equity',
    'AIFM_Infra_Core' : 'Core / Core-Plus Infrastructure',
}

_GEO = {
    'AIFM_HedgeFund' : 'Global',
    'AIFM_PrivateDebt': 'Europe (DACH, Benelux)',
    'AIFM_RealEstate' : 'Europe (France, Germany, Netherlands)',
    'AIFM_PE_Buyout'   : 'Europe (DACH, Benelux, Nordics)',
    'AIFM_Infra_Core' : 'Europe',
}

_SUBTYPE = {
    'AIFM_HedgeFund' : 'Hedge Fund — Long/Short Equity',
    'AIFM_PrivateDebt': 'Private Debt (Closed-ended)',
    'AIFM_RealEstate' : 'Real Estate (Closed-ended)',
    'AIFM_PE_Buyout'   : 'Private Equity (Closed-ended)',
    'AIFM_Infra_Core' : 'Infrastructure (Closed-ended)',
}

_LEV_LIMITS = {
    'AIFM_HedgeFund' : {'gross': 3.0, 'commitment': 2.0},
    'AIFM_PrivateDebt': {'gross': 2.0, 'commitment': 2.0},
    'AIFM_RealEstate' : {'gross': 1.5, 'commitment': 1.5},
    'AIFM_PE_Buyout'   : {'gross': 1.5, 'commitment': 1.5},
    'AIFM_Infra_Core' : {'gross': 1.5, 'commitment': 1.5},
}

_REDEMPTION = {
    'AIFM_HedgeFund' : {
        'frequency'      : 'Monthly',
        'notice_days'    : 60,
        'gate_pct'       : 10,
        'lockup_months'  : 12,
    },
    'AIFM_PrivateDebt': {
        'frequency'      : 'Quarterly',
        'notice_days'    : 90,
        'gate_pct'       : 20,
        'lockup_months'  : 24,
    },
    'AIFM_RealEstate' : {
        'frequency'      : 'Quarterly',
        'notice_days'    : 90,
        'gate_pct'       : 25,
        'lockup_months'  : 36,
    },
    'AIFM_PE_Buyout'   : {
        'frequency'      : 'Closed-ended — no periodic redemption',
        'notice_days'    : None,
        'gate_pct'       : None,
        'lockup_months'  : None,
    },
    'AIFM_Infra_Core' : {
        'frequency'      : 'Closed-ended — no periodic redemption',
        'notice_days'    : None,
        'gate_pct'       : None,
        'lockup_months'  : None,
    },
}

_FUND_TYPE = {
    'AIFM_HedgeFund' : 'hf',
    'AIFM_PrivateDebt': 'pd',
    'AIFM_RealEstate' : 're',
    'AIFM_PE_Buyout'   : 'pe',
    'AIFM_Infra_Core' : 'infra',
}

_LIQUID_FUNDS = {'AIFM_HedgeFund', 'AIFM_PrivateDebt', 'AIFM_RealEstate'}

# Simulated investor registers per fund (% of NAV holdings).
# In production these come from the transfer agent / investor register.
_INVESTOR_WEIGHTS = {
    'AIFM_HedgeFund': [
        ('HF001', 'Nordic Pension Fund',        'Pension Fund',   0.25),
        ('HF002', 'Swiss Insurance Co',          'Insurance',      0.17),
        ('HF003', 'European Family Office A',    'Family Office',  0.13),
        ('HF004', 'German Asset Manager',        'Asset Manager',  0.06),
        ('HF005', 'US Endowment Fund',           'Endowment',      0.05),
        ('HF006', 'Other investors (pooled)',    'Other',          0.34),
    ],
    'AIFM_PrivateDebt': [
        ('PD001', 'Dutch Pension Fund',          'Pension Fund',   0.38),
        ('PD002', 'German Insurance Group',      'Insurance',      0.27),
        ('PD003', 'Scandinavian Sovereign Fund', 'Sovereign',      0.18),
        ('PD004', 'European Family Office B',    'Family Office',  0.10),
        ('PD005', 'Other investors (pooled)',    'Other',          0.07),
    ],
    'AIFM_RealEstate': [
        ('RE001', 'French Pension Scheme',       'Pension Fund',   0.32),
        ('RE002', 'Belgian Insurance Co',        'Insurance',      0.24),
        ('RE003', 'UK Family Office',            'Family Office',  0.21),
        ('RE004', 'Other investors (pooled)',    'Other',          0.23),
    ],
    'AIFM_PE_Buyout': [
        ('PE001', 'Nordic Pension Fund LP',      'Pension Fund',   0.30),
        ('PE002', 'Swiss Insurance LP',          'Insurance',      0.25),
        ('PE003', 'Sovereign Wealth Fund LP',    'Sovereign',      0.20),
        ('PE004', 'European Family Office LP',   'Family Office',  0.15),
        ('PE005', 'Other LPs (pooled)',          'Other',          0.10),
    ],
    'AIFM_Infra_Core': [
        ('IF001', 'Dutch Pension Fund LP',       'Pension Fund',   0.35),
        ('IF002', 'German Insurance LP',         'Insurance',      0.30),
        ('IF003', 'Sovereign Wealth Fund LP',    'Sovereign',      0.20),
        ('IF004', 'Other LPs (pooled)',          'Other',          0.15),
    ],
}


# ── Excel style helpers ───────────────────────────────────────────────────

_BG_HEADER  = '1a1f2e'
_BG_SECTION = '0f1729'
_BG_ALT     = '1d2235'
_BG_TOTAL   = '141929'
_FG_HEADER  = 'f9fafb'
_FG_ACCENT  = '1a9ed4'
_FG_WARN    = 'ef4444'
_FG_OK      = '22c55e'
_FG_MUTED   = '9ca3af'
_FG_DIM     = '6b7280'
_SIDE       = Side(style='thin', color='374151')
_BORDER     = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)


def _hfill(c: str) -> PatternFill:
    return PatternFill('solid', fgColor=c)


def _font(bold: bool = False, color: str = _FG_HEADER, size: int = 10) -> Font:
    return Font(name='Calibri', bold=bold, color=color, size=size)


def _align(h: str = 'left', v: str = 'center') -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=False)


def _write_header(ws, row: int, col: int, value: str, width: int = 1,
                  bold: bool = True, bg: str = _BG_HEADER,
                  fg: str = _FG_HEADER) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _hfill(bg)
    cell.font = _font(bold=bold, color=fg)
    cell.border = _BORDER
    cell.alignment = _align('center')
    if width > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + width - 1)


def _write_cell(ws, row: int, col: int, value: Any, bold: bool = False,
                bg: str = _BG_SECTION, fg: str = _FG_HEADER,
                align: str = 'left', num_fmt: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _hfill(bg)
    cell.font = _font(bold=bold, color=fg)
    cell.border = _BORDER
    cell.alignment = _align(align)
    if num_fmt:
        cell.number_format = num_fmt


# ══════════════════════════════════════════════════════════════════════════
# Shared section builders
# ══════════════════════════════════════════════════════════════════════════

def _build_identification(fund_id: str, quarter: str) -> pd.DataFrame:
    """Section 1 — Fund identification. Returns DataFrame(field, value)."""
    meta = FUND_METADATA.get(fund_id, {})
    red  = _REDEMPTION[fund_id]
    lims = _LEV_LIMITS[fund_id]
    rows: list[tuple] = [
        ('FUND IDENTITY',      ''),
        ('Fund name',          meta.get('fund_name', fund_id)),
        ('Fund identifier',    fund_id),
        ('Fund LEI',           f'2138005SIM{fund_id.replace("_","")[:8]}'),
        ('Domicile',           meta.get('domicile', 'Luxembourg')),
        ('Base currency',      meta.get('currency', 'EUR')),
        ('Inception date',     meta.get('inception_date', '')),
        ('Fund type',          'Alternative Investment Fund (AIF)'),
        ('Sub-type',           _SUBTYPE[fund_id]),
        ('Strategy',           _STRATEGY[fund_id]),
        ('Geographic focus',   _GEO[fund_id]),
        ('',                   ''),
        ('AIFM',               ''),
        ('AIFM name',          _AIFM_NAME),
        ('AIFM LEI',           _AIFM_LEI),
        ('AIFM domicile',      'Luxembourg'),
        ('',                   ''),
        ('COUNTERPARTIES',     ''),
        ('Depositary',         _DEPOSITARY),
        ('Administrator',      _ADMINISTRATOR),
        ('Auditor',            _AUDITOR),
        ('',                   ''),
        ('REPORTING',          ''),
        ('Reporting period',   f'Q1 2026 (ended {quarter})'),
        ('Reporting date',     quarter),
        ('Submission date',    datetime.today().strftime('%Y-%m-%d')),
        ('Regulatory basis',   'AIFMD Art. 110 / EU 231/2013 Annex IV / ESMA v1.7 (Jul 2024)'),
        ('',                   ''),
        ('REDEMPTION TERMS',   ''),
        ('Redemption frequency', red['frequency']),
        ('Notice period (days)',
            f"{red['notice_days']} days" if red['notice_days'] else 'N/A — closed-ended'),
        ('Gate (% NAV)',
            f"{red['gate_pct']}%" if red['gate_pct'] else 'N/A — closed-ended'),
        ('Lock-up',
            f"{red['lockup_months']} months" if red['lockup_months'] else 'N/A — closed-ended'),
        ('',                   ''),
        ('LEVERAGE LIMITS (RMP)', ''),
        ('Gross method limit',      f"{lims['gross']:.1f}x NAV"),
        ('Commitment method limit', f"{lims['commitment']:.1f}x NAV"),
    ]
    return pd.DataFrame(rows, columns=['field', 'value'])


def _build_exposures(risk_df: pd.DataFrame, nav: float) -> dict[str, pd.DataFrame]:
    """Section 2 — Exposure breakdowns. Returns dict of DataFrames."""
    # asset class
    ac = (risk_df.groupby('asset_class')['market_value_eur']
          .sum().reset_index()
          .rename(columns={'market_value_eur': 'nav_eur'}))
    ac['nav_pct'] = (ac['nav_eur'] / nav * 100).round(2)
    ac = ac.sort_values('nav_eur', ascending=False).reset_index(drop=True)

    # geography
    geo_col = 'country' if 'country' in risk_df.columns else None
    if geo_col:
        geo = (risk_df.groupby(geo_col)['market_value_eur']
               .sum().reset_index()
               .rename(columns={'market_value_eur': 'nav_eur', geo_col: 'country'}))
        geo['nav_pct'] = (geo['nav_eur'] / nav * 100).round(2)
        geo = geo.sort_values('nav_eur', ascending=False).head(10).reset_index(drop=True)
    else:
        geo = pd.DataFrame(columns=['country', 'nav_eur', 'nav_pct'])

    # currency
    ccy = (risk_df.groupby('currency')['market_value_eur']
           .sum().reset_index()
           .rename(columns={'market_value_eur': 'nav_eur'}))
    ccy['nav_pct'] = (ccy['nav_eur'] / nav * 100).round(2)
    ccy = ccy.sort_values('nav_eur', ascending=False).reset_index(drop=True)

    # top 5 by absolute market value
    top5 = (risk_df.assign(abs_mv=risk_df['market_value_eur'].abs())
            .sort_values('abs_mv', ascending=False)
            .head(5)
            [['instrument_name', 'asset_class', 'sub_asset_class',
              'market_value_eur', 'currency']]
            .copy())
    top5['nav_pct'] = (top5['market_value_eur'] / nav * 100).round(2)
    top5 = top5.rename(columns={'instrument_name': 'name'}).reset_index(drop=True)
    top5.insert(0, 'rank', range(1, len(top5) + 1))

    return {
        'asset_class_breakdown': ac,
        'geography_breakdown'  : geo,
        'currency_breakdown'   : ccy,
        'top5_positions'       : top5,
    }


def _compute_gross_leverage(risk_df: pd.DataFrame,
                            nav: float) -> tuple[float, pd.DataFrame]:
    """
    EU 231/2013 Article 7: gross method.
    All exposures at abs(market value), Cash excluded.
    Returns (ratio, source_breakdown_df).
    """
    df = risk_df.copy()
    df['source'] = df.apply(
        lambda r: INSTRUMENT_SOURCE.get(
            (r['asset_class'], r.get('sub_asset_class') or ''),
            ('Cash Instrument', 'Listed'))[0],
        axis=1,
    )
    df['gross_eur'] = df['market_value_eur'].abs()
    df.loc[df['source'] == 'Excluded', 'gross_eur'] = 0.0

    by_source = (df.groupby('source')['gross_eur']
                 .sum().reset_index()
                 .query('gross_eur > 0')
                 .copy())
    by_source['pct_nav'] = (by_source['gross_eur'] / nav * 100).round(2)

    total = df['gross_eur'].sum()
    return (total / nav if nav else 0.0), by_source


def _compute_commitment_leverage(risk_df: pd.DataFrame, nav: float) -> float:
    """
    EU 231/2013 Article 8: commitment method (simplified).
    Recognises netting of long/short in the same asset class.
    Excludes Cash.
    """
    excluded = risk_df['asset_class'].isin(['Cash'])
    net = (risk_df[~excluded]
           .groupby('asset_class')['market_value_eur'].sum())
    return (net.abs().sum() / nav) if nav else 0.0


def _build_risk_measures(pnl: np.ndarray, nav: float, fund_id: str,
                         gross_lev: float, commit_lev: float,
                         liq_df: pd.DataFrame) -> pd.DataFrame:
    """Section 3 — Risk measures: VaR, leverage summary, liquidity headline."""
    pnl_w = pnl[-250:] if len(pnl) >= 250 else pnl
    v1    = var_historical(pnl_w, confidence=0.99)
    v20   = var_scale(v1, horizon=20)
    e1    = es_historical(pnl_w, confidence=0.99)

    lims = _LEV_LIMITS[fund_id]
    g_flag = '⚠ BREACH' if gross_lev  > lims['gross']      else '✓ within limit'
    c_flag = '⚠ BREACH' if commit_lev > lims['commitment'] else '✓ within limit'

    pct_1d = 0.0
    if not liq_df.empty and 'nav_pct' in liq_df.columns and 'bucket' in liq_df.columns:
        r = liq_df[liq_df['bucket'] == '1 day']
        if not r.empty:
            pct_1d = float(r['nav_pct'].iloc[0])

    rows: list[tuple] = [
        ('VaR & ES (99% confidence, historical simulation, 250 days)', ''),
        ('VaR 1-day (99%)',  f'{v1*100:.2f}%   EUR {v1*nav:,.0f}'),
        ('VaR 20-day (99%)', f'{v20*100:.2f}%   EUR {v20*nav:,.0f}'),
        ('ES 1-day (99%)',   f'{e1*100:.2f}%   EUR {e1*nav:,.0f}'),
        ('', ''),
        ('LEVERAGE', ''),
        ('Gross leverage (EU 231/2013 Art. 7)',
            f'{gross_lev:.2f}x NAV   (limit: {lims["gross"]:.1f}x)   {g_flag}'),
        ('Commitment leverage (EU 231/2013 Art. 8)',
            f'{commit_lev:.2f}x NAV   (limit: {lims["commitment"]:.1f}x)   {c_flag}'),
        ('', ''),
        ('LIQUIDITY HEADLINE', ''),
        ('% NAV liquidatable within 1 day', f'{pct_1d:.1f}%'),
    ]
    return pd.DataFrame(rows, columns=['field', 'value'])


def _build_leverage_detail(gross_lev: float, commit_lev: float, nav: float,
                           fund_id: str, breakdown: pd.DataFrame) -> pd.DataFrame:
    """Section 4 — Leverage detail: gross breakdown by source + commitment total."""
    lims = _LEV_LIMITS[fund_id]
    rows: list[tuple] = [
        ('GROSS METHOD — breakdown by source (EU 231/2013 Article 7)', '', ''),
    ]
    for _, r in breakdown.iterrows():
        rows.append((r['source'], f"EUR {r['gross_eur']:,.0f}", f"{r['pct_nav']:.2f}% NAV"))
    rows += [
        ('Total gross exposure',
            f"EUR {gross_lev*nav:,.0f}", f"{gross_lev:.2f}x NAV"),
        ('Gross leverage limit (RMP)',
            f"{lims['gross']:.1f}x NAV", ''),
        ('', '', ''),
        ('COMMITMENT METHOD (EU 231/2013 Article 8)', '', ''),
        ('Net nettable exposure',
            f"EUR {commit_lev*nav:,.0f}", f"{commit_lev:.2f}x NAV"),
        ('Commitment leverage limit (RMP)',
            f"{lims['commitment']:.1f}x NAV", ''),
    ]
    return pd.DataFrame(rows, columns=['item', 'gross_eur', 'pct_nav'])


_BUCKET_ORDER = [
    '1 day', '2-7 days', '8-30 days', '31-90 days', '91-365 days', '> 1 year',
]


def _aggregate_liquidity_buckets(liq_pos: pd.DataFrame, nav: float) -> pd.DataFrame:
    """
    Aggregate position-level liquidity_bucket assignments into an ESMA bucket summary.

    liquidity_buckets() returns a positions DataFrame — this collapses it to one
    row per bucket with NAV, % NAV, and cumulative % columns.
    """
    df = (liq_pos.groupby('liquidity_bucket', observed=True)['market_value_eur']
          .sum()
          .reindex(_BUCKET_ORDER, fill_value=0.0)
          .reset_index()
          .rename(columns={'liquidity_bucket': 'bucket', 'market_value_eur': 'nav_eur'}))
    df['nav_pct']        = (df['nav_eur'] / nav * 100).round(2) if nav else 0.0
    df['cumulative_pct'] = df['nav_pct'].cumsum().round(2)
    return df


def _build_liquidity_terms(fund_id: str, nav: float) -> pd.DataFrame:
    """Section 5 — Redemption terms + investor concentration."""
    red = _REDEMPTION[fund_id]

    inv_rows = _INVESTOR_WEIGHTS.get(fund_id, [])
    inv_df   = pd.DataFrame(inv_rows,
                            columns=['investor_id', 'investor_name',
                                     'investor_type', 'weight'])
    inv_df['aum_eur'] = inv_df['weight'] * nav

    ic     = investor_concentration(inv_df[['investor_id', 'investor_name', 'aum_eur']], nav)
    top1   = ic['largest_investor_pct'] * 100
    top3   = ic['top3_pct'] * 100
    c_flag = '⚠ Yes' if ic['concentration_flag'] else '✓ No'
    hc     = '⚠ Yes' if ic['high_concentration']  else '✓ No'

    terms_rows: list[tuple] = [
        ('Redemption frequency', red['frequency']),
        ('Notice period',
            f"{red['notice_days']} days" if red['notice_days'] else 'N/A — closed-ended'),
        ('Redemption gate',
            f"{red['gate_pct']}% of NAV" if red['gate_pct'] else 'N/A — closed-ended'),
        ('Investor lock-up',
            f"{red['lockup_months']} months" if red['lockup_months'] else 'N/A — closed-ended'),
        ('', ''),
        ('INVESTOR CONCENTRATION (ESMA thresholds)', ''),
        ('Largest single investor (% NAV)',  f'{top1:.1f}%'),
        ('Top 3 investors combined (% NAV)', f'{top3:.1f}%'),
        ('Single investor > 20% NAV flag',   c_flag),
        ('Top 3 > 50% NAV flag',             hc),
    ]
    return pd.DataFrame(terms_rows, columns=['field', 'value'])


# ══════════════════════════════════════════════════════════════════════════
# Liquid-fund builder (HF / PD / RE) — MRS-34
# ══════════════════════════════════════════════════════════════════════════

def _build_liquid(engine, fund_id: str, quarter: str) -> dict[str, pd.DataFrame]:
    risk_df   = get_risk_ready_df(engine, fund_id, _VALUATION_DATE)
    nav       = float(risk_df['market_value_eur'].sum())
    nav_hist  = query_nav_history(engine, fund_id)
    pnl       = nav_hist['pnl_pct'].dropna().values

    gross_lev, breakdown = _compute_gross_leverage(risk_df, nav)
    commit_lev           = _compute_commitment_leverage(risk_df, nav)

    liq_pos = days_to_liquidate(risk_df, pct_adv=0.25)
    liq_pos = liquidity_buckets(liq_pos)
    liq_df  = _aggregate_liquidity_buckets(liq_pos, nav)

    exposures = _build_exposures(risk_df, nav)
    terms     = _build_liquidity_terms(fund_id, nav)

    return {
        'identification'       : _build_identification(fund_id, quarter),
        'asset_class_breakdown': exposures['asset_class_breakdown'],
        'geography_breakdown'  : exposures['geography_breakdown'],
        'currency_breakdown'   : exposures['currency_breakdown'],
        'top5_positions'       : exposures['top5_positions'],
        'risk_measures'        : _build_risk_measures(
                                    pnl, nav, fund_id,
                                    gross_lev, commit_lev, liq_df),
        'leverage_detail'      : _build_leverage_detail(
                                    gross_lev, commit_lev, nav,
                                    fund_id, breakdown),
        'liquidity_buckets'    : liq_df,
        'liquidity_terms'      : terms,
        '_nav'                 : nav,
    }


# ══════════════════════════════════════════════════════════════════════════
# PE-fund builder — MRS-59
# ══════════════════════════════════════════════════════════════════════════

def _build_pe(engine, fund_id: str, quarter: str) -> dict[str, pd.DataFrame]:
    """
    Annex IV for PE buyout fund.
    Leverage at fund level only (subscription credit facility).
    Portfolio company leverage excluded per AIFMD project-finance treatment.
    AIFMD II expanded fields included.
    """
    with Session(engine) as s:
        pef = s.query(PEFund).filter_by(fund_id=fund_id).first()

        invs = (s.query(PEFundInvestment, PEPortfolioCompany)
                .join(PEPortfolioCompany,
                      PEFundInvestment.company_id == PEPortfolioCompany.company_id)
                .filter(PEFundInvestment.fund_id == fund_id)
                .all())

        nav_row = (s.query(PENavHistory)
                   .filter_by(fund_id=fund_id, company_id=None)
                   .order_by(PENavHistory.date.desc())
                   .first())

        cm = (s.query(PEFundCashManagement)
              .filter_by(fund_id=fund_id)
              .order_by(PEFundCashManagement.date.desc())
              .first())

        cfs = (s.query(PECashFlow)
               .filter_by(fund_id=fund_id)
               .all())

    nav        = float(nav_row.nav_eur) if nav_row else 0.0
    target_eur = float(pef.target_size_eur) if pef else 0.0
    drawn_eur  = sum(abs(c.amount_eur) for c in cfs if c.amount_eur < 0 and c.flow_type == 'capital_call')
    undrawn    = max(0.0, target_eur - drawn_eur)

    sub_drawn  = float(cm.sub_line_drawn)  if cm and cm.sub_line_drawn  else 0.0
    sub_limit  = float(cm.sub_line_limit)  if cm and cm.sub_line_limit  else 0.0

    # Fund-level gross leverage: (NAV + sub line drawn) / NAV
    gross_lev  = (nav + sub_drawn) / nav if nav else 1.0
    # Commitment method ≈ gross (no netting instruments for PE)
    commit_lev = gross_lev

    # multiples and IRR
    mult    = pe_multiples(engine, fund_id, as_of_date=quarter)
    irr_res = fund_irr(engine, fund_id, as_of_date=quarter)
    irr     = irr_res.get('net_irr', 0.0)

    # exposure by sector
    rows_sector: list[dict] = []
    rows_country: list[dict] = []
    rows_stage: list[dict] = []
    rows_top5: list[dict] = []
    for inv, co in invs:
        rows_sector.append({'sector': co.sector or 'Other', 'cost_eur': float(inv.cost_basis_eur)})
        rows_country.append({'country': co.country or 'Unknown', 'cost_eur': float(inv.cost_basis_eur)})
        rows_stage.append({'stage': co.investment_stage or 'Unknown', 'cost_eur': float(inv.cost_basis_eur)})
        rows_top5.append({
            'name'    : co.company_name,
            'sector'  : co.sector or '',
            'country' : co.country or '',
            'stage'   : co.investment_stage or '',
            'cost_eur': float(inv.cost_basis_eur),
        })

    def _pct_df(rows, key, label):
        df = pd.DataFrame(rows)
        df = df.groupby(key)['cost_eur'].sum().reset_index()
        total = df['cost_eur'].sum()
        df['cost_pct'] = (df['cost_eur'] / total * 100).round(2) if total else 0.0
        df = df.sort_values('cost_eur', ascending=False).reset_index(drop=True)
        df.insert(0, label, df.pop(key))
        return df

    sector_df  = _pct_df(rows_sector,  'sector',  'sector')
    country_df = _pct_df(rows_country, 'country', 'country')
    stage_df   = _pct_df(rows_stage,   'stage',   'stage')

    top5 = pd.DataFrame(rows_top5).sort_values('cost_eur', ascending=False).head(5)
    total_cost = pd.DataFrame(rows_top5)['cost_eur'].sum()
    top5['cost_pct'] = (top5['cost_eur'] / total_cost * 100).round(2) if total_cost else 0.0
    top5 = top5.reset_index(drop=True)
    top5.insert(0, 'rank', range(1, len(top5) + 1))

    lims = _LEV_LIMITS[fund_id]
    g_flag = '⚠ BREACH' if gross_lev > lims['gross'] else '✓ within limit'

    lev_rows: list[tuple] = [
        ('FUND-LEVEL LEVERAGE (EU 231/2013 Article 7)', '', ''),
        ('NAV',
            f"EUR {nav:,.0f}", ''),
        ('Subscription credit facility drawn',
            f"EUR {sub_drawn:,.0f}",
            f'{sub_drawn/nav*100:.1f}% NAV' if nav else '—'),
        ('Subscription credit facility limit',
            f"EUR {sub_limit:,.0f}", ''),
        ('Gross method leverage (fund level)',
            f'{gross_lev:.2f}x NAV',
            f'{g_flag}   limit: {lims["gross"]:.1f}x'),
        ('', '', ''),
        ('NOTE ON PROJECT-LEVEL DEBT', '', ''),
        ('Portfolio company debt',
            'Excluded from AIFMD leverage (ring-fenced at SPV level)',
            'Per EU 231/2013 Art. 7, project finance debt is excluded'),
    ]

    fund_life_remaining = None
    if pef and pef.fund_life_years and pef.vintage_year:
        fund_life_remaining = (pef.vintage_year + pef.fund_life_years) - 2026

    perf_rows: list[tuple] = [
        ('FUND PERFORMANCE', ''),
        ('Vintage year',       str(pef.vintage_year) if pef else '—'),
        ('Fund life (years)',  str(pef.fund_life_years) if pef else '—'),
        ('Fund life remaining (years)', str(fund_life_remaining) if fund_life_remaining else '—'),
        ('Strategy',           pef.strategy if pef else '—'),
        ('', ''),
        ('TARGET AND CAPITAL', ''),
        ('Target fund size',   f"EUR {target_eur:,.0f}"),
        ('Drawn capital',      f"EUR {drawn_eur:,.0f}"),
        ('Undrawn commitments',f"EUR {undrawn:,.0f}"),
        ('Fund NAV',           f"EUR {nav:,.0f}"),
        ('', ''),
        ('RETURN METRICS', ''),
        ('Gross IRR', f"{irr_res.get('gross_irr', 0)*100:.1f}%"),
        ('Net IRR',   f"{irr*100:.1f}%"),
        ('TVPI',      f"{mult.get('tvpi', 0):.2f}x"),
        ('DPI',       f"{mult.get('dpi', 0):.2f}x"),
        ('RVPI',      f"{mult.get('rvpi', 0):.2f}x"),
        ('Paid-in capital',  f"EUR {mult.get('paid_in', 0):,.0f}"),
        ('Distributions',    f"EUR {mult.get('distributions', 0):,.0f}"),
    ]

    # AIFMD II disclosures
    aifmd_ii_rows: list[tuple] = [
        ('AIFMD II — EXPANDED DISCLOSURES (Directive 2024/927/EU)', ''),
        ('', ''),
        ('LIQUIDITY MANAGEMENT TOOLS (LMTs)', ''),
        ('Suspension of redemptions',
            'Available — LP approval + board resolution required'),
        ('Side pockets',
            'Available for impaired/illiquid positions — board decision'),
        ('Capital call facility',
            f"EUR {sub_limit:,.0f} limit / EUR {sub_drawn:,.0f} drawn"),
        ('', ''),
        ('DELEGATION ARRANGEMENTS', ''),
        ('Portfolio management delegation',
            'None — managed in-house by AIFM'),
        ('Risk management delegation',
            'None — risk function internal to AIFM'),
        ('Valuation',
            'Independent appraiser — KPMG / Duff & Phelps (AIFMD Art. 19)'),
        ('', ''),
        ('PRINCIPAL MARKETS', ''),
        ('Primary market',           'Direct (bilateral) investments'),
        ('Listed instruments',        'None — fully unlisted'),
        ('FX hedging',                'None at fund level'),
        ('', ''),
        ('UNFUNDED COMMITMENTS', ''),
        ('Total unfunded (contingent leverage)',
            f"EUR {undrawn:,.0f}"),
        ('Expected drawdown next 12 months',
            'Per capital call schedule — not disclosed in Annex IV'),
    ]

    return {
        'identification'     : _build_identification(fund_id, quarter),
        'sector_exposure'    : sector_df,
        'country_exposure'   : country_df,
        'stage_exposure'     : stage_df,
        'top5_positions'     : top5,
        'leverage_detail'    : pd.DataFrame(lev_rows, columns=['item', 'value', 'note']),
        'performance'        : pd.DataFrame(perf_rows, columns=['field', 'value']),
        'aifmd_ii_disclosure': pd.DataFrame(aifmd_ii_rows, columns=['field', 'value']),
        '_nav'               : nav,
    }


# ══════════════════════════════════════════════════════════════════════════
# Infrastructure-fund builder — MRS-78
# ══════════════════════════════════════════════════════════════════════════

def _build_infra(engine, fund_id: str, quarter: str) -> dict[str, pd.DataFrame]:
    """
    Annex IV for infrastructure fund (closed-ended).
    Fund-level leverage = subscription credit line / NAV.
    Project-level debt disclosed separately (off-balance for AIFMD).
    """
    with Session(engine) as s:
        ifund = s.query(InfraFund).filter_by(fund_id=fund_id).first()

        nav_row = (s.query(InfraNavHistory)
                   .filter_by(fund_id=fund_id, asset_id=None)
                   .order_by(InfraNavHistory.date.desc())
                   .first())

        investments = (s.query(InfraFundInvestment, InfraAsset)
                       .join(InfraAsset,
                             InfraFundInvestment.asset_id == InfraAsset.asset_id)
                       .filter(InfraFundInvestment.fund_id == fund_id)
                       .all())

        # total project-level debt
        project_debt = (s.query(InfraDebt)
                        .filter(InfraDebt.asset_id.in_(
                            [inv.asset_id for inv, _ in investments]))
                        .all())

    nav           = float(nav_row.nav_eur) if nav_row else 0.0
    committed_eur = float(ifund.committed_eur) if ifund and ifund.committed_eur else 0.0
    drawn_eur     = float(ifund.drawn_eur)     if ifund and ifund.drawn_eur     else 0.0
    total_proj_debt = sum(d.outstanding_eur or 0.0 for d in project_debt)

    # fund-level leverage: no sub line for infra (typically uses committed capital only)
    gross_lev  = 1.0   # NAV / drawn equity ≈ 1x (no financial borrowing at fund level)
    commit_lev = 1.0

    # asset breakdown from infra_utils
    asset_bkd  = asset_nav_breakdown(engine, fund_id, quarter)
    sector_bkd = concentration_by_sector(engine, fund_id, quarter)
    dur_df     = duration_profile(engine, fund_id)
    infl       = inflation_sensitivity(engine, fund_id)

    # multiples and IRR
    mult = infra_multiples(engine, fund_id)
    irr  = infra_irr(engine, fund_id)

    # country and sector from investments
    rows_country: list[dict] = []
    rows_sector: list[dict] = []
    rows_top5: list[dict] = []
    for inv, asset in investments:
        nav_asset = float(asset_bkd.loc[
            asset_bkd['asset_id'] == asset.asset_id, 'nav_eur'].iloc[0]
        ) if 'asset_id' in asset_bkd.columns and asset.asset_id in asset_bkd['asset_id'].values else 0.0
        rows_country.append({'country': asset.country or 'Unknown', 'nav_eur': nav_asset})
        rows_sector.append({'sector': asset.sector or 'Other',   'nav_eur': nav_asset})
        rows_top5.append({
            'rank'    : 0,
            'name'    : asset.asset_name,
            'sector'  : asset.sector or '',
            'country' : asset.country or '',
            'nav_eur' : nav_asset,
        })

    def _pct_df(rows, key):
        df = pd.DataFrame(rows)
        df = df.groupby(key)['nav_eur'].sum().reset_index()
        total = df['nav_eur'].sum()
        df['nav_pct'] = (df['nav_eur'] / total * 100).round(2) if total else 0.0
        return df.sort_values('nav_eur', ascending=False).reset_index(drop=True)

    country_df = _pct_df(rows_country, 'country')
    # sector already in sector_bkd from concentration_by_sector

    top5 = pd.DataFrame(rows_top5).sort_values('nav_eur', ascending=False).head(5)
    top5['nav_pct'] = (top5['nav_eur'] / nav * 100).round(2) if nav else 0.0
    top5['rank'] = range(1, len(top5) + 1)
    top5 = top5[['rank', 'name', 'sector', 'country', 'nav_eur', 'nav_pct']].reset_index(drop=True)

    lims = _LEV_LIMITS[fund_id]
    lev_rows: list[tuple] = [
        ('FUND-LEVEL LEVERAGE (EU 231/2013 Article 7)', '', ''),
        ('Fund NAV',       f"EUR {nav:,.0f}", ''),
        ('Committed capital', f"EUR {committed_eur:,.0f}", ''),
        ('Drawn capital',  f"EUR {drawn_eur:,.0f}", ''),
        ('Fund-level debt', 'None — no subscription credit line', ''),
        ('Gross method leverage (fund level)', f'{gross_lev:.2f}x NAV',
            f'✓ within limit ({lims["gross"]:.1f}x)'),
        ('Commitment method leverage',         f'{commit_lev:.2f}x NAV',
            f'✓ within limit ({lims["commitment"]:.1f}x)'),
        ('', '', ''),
        ('PROJECT-LEVEL DEBT (off-balance, disclosed separately)', '', ''),
        ('Total project-level debt outstanding',
            f"EUR {total_proj_debt:,.0f}",
            'Ring-fenced at SPV level — excluded from AIFMD leverage'),
        ('Number of debt-bearing assets',
            str(len([d for d in project_debt if d.outstanding_eur])),
            'Project finance structure'),
        ('AIFMD treatment',
            'Project debt excluded from Art. 7 gross leverage',
            'EU 231/2013 Art. 7 — project finance carve-out'),
    ]

    fund_life_remaining = None
    if ifund and ifund.fund_life_years and ifund.vintage_year:
        fund_life_remaining = (ifund.vintage_year + ifund.fund_life_years) - 2026

    wad = dur_df.attrs.get('weighted_avg_remaining_years', None)

    perf_rows: list[tuple] = [
        ('FUND PERFORMANCE', ''),
        ('Vintage year',     str(ifund.vintage_year) if ifund else '—'),
        ('Fund life (years)', str(ifund.fund_life_years) if ifund else '—'),
        ('Fund life remaining (years)', str(fund_life_remaining) if fund_life_remaining else '—'),
        ('Classification',   ifund.aifmd_classification if ifund else '—'),
        ('', ''),
        ('CAPITAL', ''),
        ('Committed capital', f"EUR {committed_eur:,.0f}"),
        ('Drawn capital',     f"EUR {drawn_eur:,.0f}"),
        ('Fund NAV',          f"EUR {nav:,.0f}"),
        ('', ''),
        ('RETURN METRICS', ''),
        ('Net IRR',   f"{irr*100:.1f}%" if irr else '—'),
        ('MOIC',      f"{mult.get('moic', 0):.2f}x"),
        ('DPI',       f"{mult.get('dpi', 0):.2f}x"),
        ('RVPI',      f"{mult.get('rvpi', 0):.2f}x"),
        ('', ''),
        ('INFRASTRUCTURE CHARACTERISTICS', ''),
        ('Number of assets',              str(len(investments))),
        ('Weighted avg concession life (years)',
            f"{wad:.1f}" if wad else '—'),
        ('Weighted avg inflation linkage',
            f"{infl.get('weighted_avg_linkage', 0)*100:.1f}%"),
        ('% fully inflation-linked',
            f"{infl.get('pct_fully_linked', 0):.1f}%"),
        ('% partially inflation-linked',
            f"{infl.get('pct_partially_linked', 0):.1f}%"),
    ]

    return {
        'identification'  : _build_identification(fund_id, quarter),
        'asset_breakdown' : asset_bkd,
        'sector_breakdown': sector_bkd,
        'country_breakdown': country_df,
        'top5_positions'  : top5,
        'leverage_detail' : pd.DataFrame(lev_rows, columns=['item', 'value', 'note']),
        'performance'     : pd.DataFrame(perf_rows, columns=['field', 'value']),
        '_nav'            : nav,
    }


# ══════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════

_BUILDERS = {
    'hf'   : _build_liquid,
    'pd'   : _build_liquid,
    're'   : _build_liquid,
    'pe'   : _build_pe,
    'infra': _build_infra,
}


def build_annex_iv(engine, fund_id: str,
                   quarter: str = '2026-03-31') -> dict[str, pd.DataFrame]:
    """
    Build Annex IV report for a single fund.

    Parameters
    ----------
    engine : SQLAlchemy Engine
    fund_id : str
        One of: AIFM_HedgeFund, AIFM_PrivateDebt, AIFM_RealEstate,
                AIFM_PE_Buyout, AIFM_Infra_Core
    quarter : str
        Reporting period end date, e.g. '2026-03-31'

    Returns
    -------
    dict[str, pd.DataFrame]
        Keys depend on fund type. Always present:
          identification, leverage_detail, performance (PE/infra)
          or risk_measures / liquidity_buckets (liquid funds).
    """
    ftype = _FUND_TYPE.get(fund_id)
    if ftype is None:
        raise ValueError(f'Unknown fund_id: {fund_id}')
    return _BUILDERS[ftype](engine, fund_id, quarter)


# ══════════════════════════════════════════════════════════════════════════
# Excel export
# ══════════════════════════════════════════════════════════════════════════

_EXPORT_FUNDS = [
    'AIFM_HedgeFund',
    'AIFM_PrivateDebt',
    'AIFM_RealEstate',
    'AIFM_PE_Buyout',
    'AIFM_Infra_Core',
]

_SHEET_NAMES = {
    'AIFM_HedgeFund' : 'HedgeFund',
    'AIFM_PrivateDebt': 'PrivateDebt',
    'AIFM_RealEstate' : 'RealEstate',
    'AIFM_PE_Buyout'   : 'PEBuyout',
    'AIFM_Infra_Core' : 'InfraCore',
}


def _write_section_header(ws, row: int, label: str, ncols: int = 3) -> int:
    _write_header(ws, row, 1, label, width=ncols, bg=_BG_HEADER)
    return row + 1


def _write_field_value_df(ws, row: int, df: pd.DataFrame,
                          ncols: int = 2) -> int:
    for idx, r in df.iterrows():
        f, v = str(r.iloc[0]), str(r.iloc[1]) if len(r) > 1 else ''
        if f == '' and v == '':
            row += 1
            continue
        is_header = (v == '' and f == f.upper() and f != '')
        bg = _BG_HEADER if is_header else (_BG_ALT if idx % 2 == 0 else _BG_SECTION)
        fg = _FG_ACCENT if is_header else _FG_HEADER
        _write_cell(ws, row, 1, f, bold=is_header, bg=bg, fg=fg)
        if ncols >= 2:
            _write_cell(ws, row, 2, v, bg=bg, fg=_FG_HEADER)
        row += 1
    return row + 1


def _write_df_table(ws, row: int, df: pd.DataFrame,
                    title: str = '', alt: bool = True) -> int:
    if title:
        _write_header(ws, row, 1, title,
                      width=len(df.columns), bg=_BG_HEADER)
        row += 1
    # column headers
    for ci, col in enumerate(df.columns, start=1):
        _write_header(ws, row, ci, str(col), width=1, bg='262d40')
    row += 1
    for idx, r in df.iterrows():
        bg = _BG_ALT if (alt and idx % 2 == 0) else _BG_SECTION
        for ci, val in enumerate(r, start=1):
            _write_cell(ws, row, ci, val, bg=bg)
        row += 1
    return row + 1


def _write_fund_sheet(wb: Workbook, fund_id: str,
                      rpt: dict[str, pd.DataFrame],
                      quarter: str) -> None:
    sheet = _SHEET_NAMES[fund_id]
    ws = wb.create_sheet(title=sheet)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14

    nav = rpt.get('_nav', 0.0)
    meta = FUND_METADATA.get(fund_id, {})

    # Title
    _write_header(ws, 1, 1,
                  f"Annex IV Regulatory Report — {meta.get('fund_name', fund_id)}",
                  width=5, bg=_BG_HEADER)
    _write_header(ws, 2, 1,
                  f"Fund: {fund_id}   |   NAV: EUR {nav/1e6:,.1f}M   "
                  f"|   Reporting period: Q1 2026 ({quarter})   "
                  f"|   Generated: {datetime.today().strftime('%Y-%m-%d')}",
                  width=5, bg=_BG_SECTION, bold=False, fg=_FG_MUTED)
    _write_header(ws, 3, 1,
                  'Regulatory basis: AIFMD Art. 110 / EU 231/2013 Annex IV / '
                  'ESMA technical guidance v1.7 (July 2024)',
                  width=5, bg=_BG_SECTION, bold=False, fg=_FG_DIM)

    row = 5

    # Section 1 — Identification
    row = _write_field_value_df(ws, row, rpt['identification'])

    # Sections depend on fund type
    ftype = _FUND_TYPE[fund_id]

    if ftype in ('hf', 'pd', 're'):
        # Section 2 — Exposures
        row = _write_df_table(ws, row, rpt['asset_class_breakdown'],
                              'Section 2.1 — Asset class breakdown')
        row = _write_df_table(ws, row, rpt['geography_breakdown'],
                              'Section 2.2 — Geographic breakdown (top 10)')
        row = _write_df_table(ws, row, rpt['currency_breakdown'],
                              'Section 2.3 — Currency breakdown')
        row = _write_df_table(ws, row, rpt['top5_positions'],
                              'Section 2.4 — Top 5 positions by absolute exposure')
        # Section 3 — Risk measures
        row = _write_field_value_df(ws, row, rpt['risk_measures'])
        # Section 4 — Leverage
        row = _write_df_table(ws, row, rpt['leverage_detail'],
                              'Section 4 — Leverage detail')
        # Section 5 — Liquidity
        row = _write_df_table(ws, row, rpt['liquidity_buckets'],
                              'Section 5.1 — Liquidity profile (ESMA buckets)')
        row = _write_field_value_df(ws, row, rpt['liquidity_terms'])

    elif ftype == 'pe':
        row = _write_df_table(ws, row, rpt['sector_exposure'],
                              'Section 2.1 — Portfolio exposure by sector (cost basis)')
        row = _write_df_table(ws, row, rpt['country_exposure'],
                              'Section 2.2 — Portfolio exposure by country')
        row = _write_df_table(ws, row, rpt['stage_exposure'],
                              'Section 2.3 — Portfolio exposure by investment stage')
        row = _write_df_table(ws, row, rpt['top5_positions'],
                              'Section 2.4 — Top 5 investments')
        row = _write_df_table(ws, row, rpt['leverage_detail'],
                              'Section 3 — Leverage (fund-level)')
        row = _write_field_value_df(ws, row, rpt['performance'])
        row = _write_field_value_df(ws, row, rpt['aifmd_ii_disclosure'])

    elif ftype == 'infra':
        row = _write_df_table(ws, row, rpt['asset_breakdown'],
                              'Section 2.1 — Asset NAV breakdown')
        row = _write_df_table(ws, row, rpt['sector_breakdown'],
                              'Section 2.2 — Sector concentration')
        row = _write_df_table(ws, row, rpt['country_breakdown'],
                              'Section 2.3 — Geographic breakdown')
        row = _write_df_table(ws, row, rpt['top5_positions'],
                              'Section 2.4 — Top 5 assets by NAV')
        row = _write_df_table(ws, row, rpt['leverage_detail'],
                              'Section 3 — Leverage')
        row = _write_field_value_df(ws, row, rpt['performance'])

    ws.freeze_panes = 'A5'


def _write_summary_sheet(wb: Workbook,
                         reports: dict[str, dict[str, pd.DataFrame]],
                         quarter: str) -> None:
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    funds = list(reports.keys())
    nf    = len(funds)

    _write_header(ws, 1, 1,
                  'Annex IV Regulatory Report — Cross-Fund Summary',
                  width=1 + nf, bg=_BG_HEADER)
    _write_header(ws, 2, 1,
                  f'Reporting period: Q1 2026 ({quarter})   |   '
                  f'Generated: {datetime.today().strftime("%Y-%m-%d")}   |   '
                  'Regulatory basis: AIFMD Art. 110 / EU 231/2013 Annex IV',
                  width=1 + nf, bg=_BG_SECTION, bold=False, fg=_FG_MUTED)

    row = 4
    _write_header(ws, row, 1, 'Key metric', bg=_BG_HEADER)
    for i, fid in enumerate(funds, start=2):
        meta = FUND_METADATA.get(fid, {})
        _write_header(ws, row, i, meta.get('fund_name', fid), bg=_BG_HEADER)

    summary_rows: list[tuple] = [
        ('NAV (EUR M)',),
        ('Strategy',),
        ('Geography',),
        ('Gross leverage',),
        ('Commitment leverage',),
        ('VaR 1-day 99% (%)',),
        ('VaR 20-day 99% (%)',),
        ('Redemption frequency',),
        ('Fund life remaining (years)',),
    ]

    row = 5
    for si, (label,) in enumerate(summary_rows):
        bg = _BG_ALT if si % 2 == 0 else _BG_SECTION
        _write_cell(ws, row, 1, label, bold=True, bg=bg, align='left')
        for i, fid in enumerate(funds, start=2):
            rpt   = reports[fid]
            ftype = _FUND_TYPE[fid]
            nav   = rpt.get('_nav', 0.0)
            val   = '—'

            if label == 'NAV (EUR M)':
                val = f"EUR {nav/1e6:,.1f}M"
            elif label == 'Strategy':
                val = _STRATEGY.get(fid, '—')
            elif label == 'Geography':
                val = _GEO.get(fid, '—')
            elif label == 'Gross leverage':
                if ftype in ('hf', 'pd', 're'):
                    risk_df = None  # already built
                    lev_df  = rpt.get('leverage_detail', pd.DataFrame())
                    lev_row = lev_df[lev_df['item'].str.startswith('Total gross')
                                     if 'item' in lev_df.columns else
                                     lev_df.index < 0]
                    if not lev_row.empty and 'pct_nav' in lev_row.columns:
                        val = str(lev_row['pct_nav'].iloc[0])
                    else:
                        val = '—'
                else:
                    lev_df = rpt.get('leverage_detail', pd.DataFrame())
                    if 'item' in lev_df.columns:
                        r = lev_df[lev_df['item'].str.contains('Gross method leverage', na=False)]
                        val = str(r['value'].iloc[0]) if not r.empty else '—'
            elif label == 'Commitment leverage':
                if ftype in ('hf', 'pd', 're'):
                    lev_df = rpt.get('leverage_detail', pd.DataFrame())
                    if 'item' in lev_df.columns:
                        r = lev_df[lev_df['item'].str.contains('Net nettable', na=False)]
                        val = str(r['pct_nav'].iloc[0]) if not r.empty else '—'
                else:
                    lev_df = rpt.get('leverage_detail', pd.DataFrame())
                    if 'item' in lev_df.columns:
                        r = lev_df[lev_df['item'].str.contains('Commitment method', na=False)]
                        val = str(r['value'].iloc[0]) if not r.empty else '—'
            elif label in ('VaR 1-day 99% (%)', 'VaR 20-day 99% (%)'):
                if ftype in ('hf', 'pd', 're'):
                    rm = rpt.get('risk_measures', pd.DataFrame())
                    if 'field' in rm.columns:
                        key = 'VaR 1-day (99%)' if '1-day' in label else 'VaR 20-day (99%)'
                        r   = rm[rm['field'] == key]
                        val = str(r['value'].iloc[0]).split()[0] if not r.empty else '—'
                else:
                    val = 'N/A — illiquid'
            elif label == 'Redemption frequency':
                val = _REDEMPTION.get(fid, {}).get('frequency', '—')
            elif label == 'Fund life remaining (years)':
                if ftype in ('pe', 'infra'):
                    perf = rpt.get('performance', pd.DataFrame())
                    if 'field' in perf.columns:
                        r = perf[perf['field'] == 'Fund life remaining (years)']
                        val = str(r['value'].iloc[0]) if not r.empty else '—'
                else:
                    val = 'Open-ended'

            _write_cell(ws, row, i, val, bg=bg, align='right')
        row += 1

    ws.freeze_panes = 'A5'


def export_annex_iv_excel(
    engine=None,
    quarter: str = '2026-03-31',
    output_dir: str = 'data',
    fund_ids: list[str] | None = None,
) -> str:
    """
    Build Annex IV reports for all (or specified) AIFM funds and write
    a CSSF-ready Excel workbook.

    Parameters
    ----------
    engine : SQLAlchemy Engine, optional
    quarter : str
        Reporting period end date, e.g. '2026-03-31'.
    output_dir : str
        Directory for output file.
    fund_ids : list[str], optional
        Funds to include. Defaults to all five AIFM funds.

    Returns
    -------
    str
        Full path to the written workbook.
    """
    if engine is None:
        engine = get_engine()
    if fund_ids is None:
        fund_ids = _EXPORT_FUNDS

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, f'annex_iv_report_{quarter}.xlsx')

    print(f'Annex IV export — reporting period {quarter}')

    reports: dict[str, dict] = {}
    for fid in fund_ids:
        print(f'  Building {fid}...', end=' ', flush=True)
        rpt = build_annex_iv(engine, fid, quarter)
        reports[fid] = rpt
        nav = rpt.get('_nav', 0.0)
        print(f'NAV EUR {nav/1e6:,.1f}M')

    wb = Workbook()
    _write_summary_sheet(wb, reports, quarter)
    for fid in fund_ids:
        _write_fund_sheet(wb, fid, reports[fid], quarter)

    wb.save(out_path)
    print(f'\nWritten: {out_path}')
    return out_path


if __name__ == '__main__':
    export_annex_iv_excel()
