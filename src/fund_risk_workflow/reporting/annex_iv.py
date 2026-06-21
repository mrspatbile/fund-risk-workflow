"""
src/annex_iv.py
===============
Annex IV regulatory transparency report — AIFMD Art. 110 / EU231/2013.
Quarterly submission to the CSSF for all AIFM funds.

Usage
-----
    from fund_risk_workflow.reporting.annex_iv import build_annex_iv, export_annex_iv_excel

    rpt  = build_annex_iv(engine, 'AIFM_HedgeFund', quarter='2026-03-31')
    rpt['identification']
    rpt['asset_class_breakdown']
    rpt['risk_measures']
    rpt['leverage_detail']
    rpt['liquidity_buckets']

    path = export_annex_iv_excel(engine, quarter='2026-03-31')

Output (Excel)
--------------
    data/annex_iv_report_<quarter>.xlsx

    Sheets: Summary | AIFM_HedgeFund | AIFM_PrivateDebt | AIFM_RealEstate
            | AIFM_PE_Buyout | AIFM_Infra_Core

Regulatory basis
----------------
    AIFMD Art. 110 — Annex IV transparency reporting
    EU231/2013 Articles 110-121 — AIFM regulatory reporting requirements
    ESMA technical guidance v1.7 (July 2024) — Annex IV field definitions
    AIFMD II (Directive 2024/927/EU) — expanded LMT and delegation disclosures
"""


import os
from datetime import datetime
from typing import Any

import numpy as np
import pandas as pd
import sqlalchemy as sa
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from fund_risk_workflow.data.database import (
    InfraAsset,
    InfraDebt,
    InfraFund,
    InfraFundInvestment,
    InfraNavHistory,
    PECashFlow,
    PEFund,
    PEFundCashManagement,
    PEFundInvestment,
    PENavHistory,
    PEPortfolioCompany,
    get_engine,
    query_nav_history,
)
from fund_risk_workflow.config import LIQUIDITY_BUCKET_ORDER, VALUATION_DATE
from fund_risk_workflow.data.enrichment import get_risk_ready_df
from fund_risk_workflow.risk.infra_utils import (
    asset_nav_breakdown,
    concentration_by_sector,
    duration_profile,
    inflation_sensitivity,
    infra_irr,
    infra_multiples,
)
from fund_risk_workflow.risk.leverage_config import INSTRUMENT_SOURCE
from fund_risk_workflow.risk.pe_utils import fund_irr, pe_multiples
from fund_risk_workflow.risk.risk_utils import (
    compute_liquidity_profile,
    es_historical,
    investor_concentration,
    var_historical,
    var_scale,
)


# ══════════════════════════════════════════════════════════════════════════
# Static configuration
# ══════════════════════════════════════════════════════════════════════════

_VALUATION_DATE = VALUATION_DATE
_AIFM_NAME      = 'ManCo SA'
_AIFM_LEI       = '213800SIMULATED0001'
_DEPOSITARY     = 'BNP Paribas Securities Services Luxembourg'
_ADMINISTRATOR  = 'IQ-EQ Luxembourg'
_AUDITOR        = 'PricewaterhouseCoopers Luxembourg'

_STRATEGY = {
    'AIFM_HedgeFund'  : 'Long/Short Equity',
    'AIFM_PrivateDebt': 'Private Debt — Direct Lending',
    'AIFM_RealEstate' : 'Core Real Estate',
    'AIFM_PE_Buyout'  : 'Leveraged Buyout / Growth Equity',
    'AIFM_Infra_Core' : 'Core / Core-Plus Infrastructure',
}

_GEO = {
    'AIFM_HedgeFund'  : 'Global',
    'AIFM_PrivateDebt': 'Europe (DACH, Benelux)',
    'AIFM_RealEstate' : 'Europe (France, Germany, Netherlands)',
    'AIFM_PE_Buyout'  : 'Europe (DACH, Benelux, Nordics)',
    'AIFM_Infra_Core' : 'Europe',
}

_SUBTYPE = {
    'AIFM_HedgeFund'  : 'Hedge Fund — Long/Short Equity',
    'AIFM_PrivateDebt': 'Private Debt (Closed-ended)',
    'AIFM_RealEstate' : 'Real Estate (Closed-ended)',
    'AIFM_PE_Buyout'  : 'Private Equity (Closed-ended)',
    'AIFM_Infra_Core' : 'Infrastructure (Closed-ended)',
}

_LEV_LIMITS = {
    'AIFM_HedgeFund'  : {'gross': 3.0, 'commitment': 2.0},
    'AIFM_PrivateDebt': {'gross': 2.0, 'commitment': 2.0},
    'AIFM_RealEstate' : {'gross': 1.5, 'commitment': 1.5},
    'AIFM_PE_Buyout'  : {'gross': 1.5, 'commitment': 1.5},
    'AIFM_Infra_Core' : {'gross': 1.5, 'commitment': 1.5},
}

_REDEMPTION = {
    'AIFM_HedgeFund'  : {'frequency': 'Daily',      'notice_days': 5,   'gate_pct': 10,  'lockup_months': 12},
    'AIFM_PrivateDebt': {'frequency': 'Quarterly',   'notice_days': 90,  'gate_pct': 20,  'lockup_months': 24},
    'AIFM_RealEstate' : {'frequency': 'Quarterly',   'notice_days': 90,  'gate_pct': 25,  'lockup_months': 36},
    'AIFM_PE_Buyout'  : {'frequency': 'Closed-ended — no periodic redemption', 'notice_days': None, 'gate_pct': None, 'lockup_months': None},
    'AIFM_Infra_Core' : {'frequency': 'Closed-ended — no periodic redemption', 'notice_days': None, 'gate_pct': None, 'lockup_months': None},
}

_FUND_TYPE = {
    'AIFM_HedgeFund'  : 'hf',
    'AIFM_PrivateDebt': 'pd',
    'AIFM_RealEstate' : 're',
    'AIFM_PE_Buyout'  : 'pe',
    'AIFM_Infra_Core' : 'infra',
}

_INVESTOR_WEIGHTS = {
    'AIFM_HedgeFund': [
        ('HF001', 'Nordic Pension Fund',        'Pension Fund',  0.25),
        ('HF002', 'Swiss Insurance Co',          'Insurance',     0.17),
        ('HF003', 'European Family Office A',    'Family Office', 0.13),
        ('HF004', 'German Asset Manager',        'Asset Manager', 0.06),
        ('HF005', 'US Endowment Fund',           'Endowment',     0.05),
        ('HF006', 'Other investors (pooled)',    'Other',         0.34),
    ],
    'AIFM_PrivateDebt': [
        ('PD001', 'Dutch Pension Fund',          'Pension Fund',  0.38),
        ('PD002', 'German Insurance Group',      'Insurance',     0.27),
        ('PD003', 'Scandinavian Sovereign Fund', 'Sovereign',     0.18),
        ('PD004', 'European Family Office B',    'Family Office', 0.10),
        ('PD005', 'Other investors (pooled)',    'Other',         0.07),
    ],
    'AIFM_RealEstate': [
        ('RE001', 'French Pension Scheme',       'Pension Fund',  0.32),
        ('RE002', 'Belgian Insurance Co',        'Insurance',     0.24),
        ('RE003', 'UK Family Office',            'Family Office', 0.21),
        ('RE004', 'Other investors (pooled)',    'Other',         0.23),
    ],
    'AIFM_PE_Buyout': [
        ('PE001', 'Nordic Pension Fund LP',      'Pension Fund',  0.30),
        ('PE002', 'Swiss Insurance LP',          'Insurance',     0.25),
        ('PE003', 'Sovereign Wealth Fund LP',    'Sovereign',     0.20),
        ('PE004', 'European Family Office LP',   'Family Office', 0.15),
        ('PE005', 'Other LPs (pooled)',          'Other',         0.10),
    ],
    'AIFM_Infra_Core': [
        ('IF001', 'Dutch Pension Fund LP',       'Pension Fund',  0.35),
        ('IF002', 'German Insurance LP',         'Insurance',     0.30),
        ('IF003', 'Sovereign Wealth Fund LP',    'Sovereign',     0.20),
        ('IF004', 'Other LPs (pooled)',          'Other',         0.15),
    ],
}


# ══════════════════════════════════════════════════════════════════════════
# Shared row-builder helpers
# ══════════════════════════════════════════════════════════════════════════

def _eur(v: float) -> str:
    return f"EUR {v:,.0f}"


def _pct_df(rows: list[dict], key: str, value_col: str = 'nav_eur') -> pd.DataFrame:
    """Aggregate a list of dicts by key, compute nav_pct, sort descending."""
    df    = pd.DataFrame(rows).groupby(key)[value_col].sum().reset_index()
    total = df[value_col].sum()
    df['nav_pct'] = (df[value_col] / total * 100).round(2) if total else 0.0
    df = df.sort_values(value_col, ascending=False).reset_index(drop=True)
    df[value_col] = df[value_col].map('{:,.0f}'.format)

    return df


def _fund_life_remaining(fund_obj) -> int | None:
    if fund_obj and fund_obj.fund_life_years and fund_obj.vintage_year:
        return (fund_obj.vintage_year + fund_obj.fund_life_years) - 2026
    return None


def _fmt_red(red: dict, field: str) -> str:
    v = red.get(field)
    if v is None:
        return 'N/A — closed-ended'
    if field == 'notice_days':
        return f"{v} days"
    if field == 'gate_pct':
        return f"{v}%"
    if field == 'lockup_months':
        return f"{v} months"
    return str(v)


def _capital_rows(committed: float, drawn: float, nav: float,
                  undrawn: float | None = None) -> list[tuple]:
    rows = [
        ('CAPITAL', ''),
        ('Committed capital', _eur(committed)),
        ('Drawn capital',     _eur(drawn)),
    ]
    if undrawn is not None:
        rows.append(('Undrawn commitments', _eur(undrawn)))
    rows.append(('Fund NAV', _eur(nav)))
    return rows


def _return_rows(mult: dict, irr: float,
                 gross_irr: float | None = None,
                 moic_label: str = 'MOIC') -> list[tuple]:   # <-- added
    rows: list[tuple] = [('RETURN METRICS', '')]
    if gross_irr is not None:
        rows.append(('Gross IRR', f"{gross_irr*100:.1f}%"))
    rows += [
        ('Net IRR',   f"{irr*100:.1f}%" if irr else '—'),
        (moic_label,  f"{mult.get('tvpi', mult.get('moic', 0)):.2f}x"),  # <-- uses label
        ('DPI',       f"{mult.get('dpi', 0):.2f}x"),
        ('RVPI',      f"{mult.get('rvpi', 0):.2f}x"),
    ]
    return rows


def _fund_header_rows(fund_obj, life_remaining: int | None,
                      extra: list[tuple] | None = None) -> list[tuple]:
    rows = [
        ('FUND PERFORMANCE', ''),
        ('Vintage year',                str(fund_obj.vintage_year)    if fund_obj else '—'),
        ('Fund life (years)',           str(fund_obj.fund_life_years) if fund_obj else '—'),
        ('Fund life remaining (years)', str(life_remaining)           if life_remaining else '—'),
    ]
    if extra:
        rows += extra
    return rows


def _lev_flag(lev: float, fund_id: str, method: str = 'gross') -> str:
    limit = _LEV_LIMITS[fund_id][method]
    flag  = '⚠ BREACH' if lev > limit else '✓ within limit'
    return f"{lev:.2f}x NAV   (limit: {limit:.1f}x)   {flag}"


def _project_debt_note() -> list[tuple]:
    return [
        ('', '', ''),
        ('NOTE ON PROJECT-LEVEL DEBT', '', ''),
        ('Portfolio company debt',
            'Excluded from AIFMD leverage (ring-fenced at SPV level)',
            'Per EU231/2013 Art. 7, project finance debt is excluded'),
    ]


def _get_fund_metadata(engine: sa.Engine, fund_id: str) -> dict:
    """Get fund metadata from the funds table."""
    query = text(
        """
        SELECT fund_name, domicile, currency, inception_date
        FROM funds
        WHERE fund_id = :fund_id
        """
    )

    with engine.connect() as conn:
        result = conn.execute(query, {"fund_id": fund_id}).mappings().first()

    if not result:
        return {}

    return {
        "fund_name": result["fund_name"],
        "domicile": result["domicile"],
        "currency": result["currency"],
        "inception_date": result["inception_date"],
    }


# ══════════════════════════════════════════════════════════════════════════
# Section builders shared across fund types
# ══════════════════════════════════════════════════════════════════════════

def _build_identification(engine: sa.Engine, fund_id: str, quarter: str) -> pd.DataFrame:
    meta = _get_fund_metadata(engine, fund_id)
    red  = _REDEMPTION[fund_id]
    lims = _LEV_LIMITS[fund_id]
    rows = [
        ('FUND IDENTITY',      ''),
        ('Fund name',          meta.get('fund_name', fund_id)),
        ('Fund identifier',    fund_id),
        ('Fund LEI',           f'2138005SIM{fund_id.replace("_","")[:8]}'),
        ('Domicile',           meta.get('domicile', 'Luxembourg')),
        ('Base currency',      meta.get('currency', 'EUR')),
        ('Inception date',     meta.get('inception_date', '')),
        ('Fund type',          'Alternative Investment Fund (AIF)'),
        ('Sub-type',           _SUBTYPE[fund_id]),
        ('Strategy',           _STRATEGY[fund_id]),
        ('Geographic focus',   _GEO[fund_id]),
        ('',                   ''),
        ('AIFM',               ''),
        ('AIFM name',          _AIFM_NAME),
        ('AIFM LEI',           _AIFM_LEI),
        ('AIFM domicile',      'Luxembourg'),
        ('',                   ''),
        ('COUNTERPARTIES',     ''),
        ('Depositary',         _DEPOSITARY),
        ('Administrator',      _ADMINISTRATOR),
        ('Auditor',            _AUDITOR),
        ('',                   ''),
        ('REPORTING',          ''),
        ('Reporting period',   f'Q1 2026 (ended {quarter})'),
        ('Reference date',     quarter),
        ('Filing date',        (pd.Timestamp(quarter) + pd.Timedelta(days=15)).strftime('%Y-%m-%d')),
        ('Regulatory basis',   'AIFMD Art. 110 / EU231/2013 Annex IV / ESMA v1.7 (Jul 2024)'),
        ('',                   ''),
        ('REDEMPTION TERMS',   ''),
        ('Redemption frequency', red['frequency']),
        ('Notice period (days)', _fmt_red(red, 'notice_days')),
        ('Gate (% NAV)',          _fmt_red(red, 'gate_pct')),
        ('Lock-up',               _fmt_red(red, 'lockup_months')),
        ('',                      ''),
        ('LEVERAGE LIMITS (RMP)', ''),
        ('Gross method limit',      f"{lims['gross']:.1f}x NAV"),
        ('Commitment method limit', f"{lims['commitment']:.1f}x NAV"),
    ]
    return pd.DataFrame(rows, columns=['field', 'value'])


def _build_exposures(risk_df: pd.DataFrame, nav: float) -> dict[str, pd.DataFrame]:
    def _agg(group_col, rename=None):
        col = rename or group_col
        df  = (risk_df.groupby(group_col)['market_value_eur']
               .sum().reset_index()
               .rename(columns={'market_value_eur': 'nav_eur', group_col: col}))
        df['nav_pct'] = (df['nav_eur'] / nav * 100).round(2)
        return df.sort_values('nav_eur', ascending=False).reset_index(drop=True)

    ac  = _agg('asset_class')
    ccy = _agg('currency')
    geo = (_agg('country').head(10)
           if 'country' in risk_df.columns
           else pd.DataFrame(columns=['country', 'nav_eur', 'nav_pct']))

    top5 = (risk_df.assign(abs_mv=risk_df['market_value_eur'].abs())
            .sort_values('abs_mv', ascending=False).head(5)
            [['instrument_name', 'asset_class', 'sub_asset_class',
              'market_value_eur', 'currency']].copy())
    top5['nav_pct'] = (top5['market_value_eur'] / nav * 100).round(2)
    top5 = top5.rename(columns={'instrument_name': 'name'}).reset_index(drop=True)
    top5.insert(0, 'rank', range(1, len(top5) + 1))

    return {
        'asset_class_breakdown': ac,
        'geography_breakdown'  : geo,
        'currency_breakdown'   : ccy,
        'top5_positions'       : top5,
    }


def _compute_gross_leverage(risk_df: pd.DataFrame,
                            nav: float) -> tuple[float, pd.DataFrame]:
    df = risk_df.copy()
    df['source'] = df.apply(
        lambda r: INSTRUMENT_SOURCE.get(
            (r['asset_class'], r.get('sub_asset_class') or ''),
            ('Cash Instrument', 'Listed'))[0],
        axis=1,
    )
    df['gross_eur'] = df['market_value_eur'].abs()
    df.loc[df['source'] == 'Excluded', 'gross_eur'] = 0.0

    by_source = (df.groupby('source')['gross_eur']
                 .sum().reset_index()
                 .query('gross_eur > 0').copy())
    by_source['pct_nav'] = (by_source['gross_eur'] / nav * 100).round(2)

    total = df['gross_eur'].sum()
    return (total / nav if nav else 0.0), by_source


def _compute_commitment_leverage(risk_df: pd.DataFrame, nav: float) -> float:
    net = (risk_df[~risk_df['asset_class'].isin(['Cash'])]
           .groupby('asset_class')['market_value_eur'].sum())
    return (net.abs().sum() / nav) if nav else 0.0


def _build_risk_measures(pnl: np.ndarray, nav: float, fund_id: str,
                         gross_lev: float, commit_lev: float,
                         liq_df: pd.DataFrame) -> pd.DataFrame:
    pnl_w = pnl[-250:] if len(pnl) >= 250 else pnl
    v1    = var_historical(pnl_w, confidence=0.99)
    v20   = var_scale(v1, horizon=20)
    e1    = es_historical(pnl_w, confidence=0.99)

    pct_1d = 0.0
    if not liq_df.empty and 'nav_pct' in liq_df.columns:
        r = liq_df[liq_df['bucket'] == '1 day']
        if not r.empty:
            pct_1d = float(r['nav_pct'].iloc[0])

    rows = [
        ('VaR & ES (99%, historical, 250 days)', ''),
        ('VaR 1-day (99%)',  f'{v1*100:.2f}%   {_eur(v1*nav)}'),
        ('VaR 20-day (99%)', f'{v20*100:.2f}%   {_eur(v20*nav)}'),
        ('ES 1-day (99%)',   f'{e1*100:.2f}%   {_eur(e1*nav)}'),
        ('', ''),
        ('LEVERAGE', ''),

        ('Gross leverage (EU231/2013 Art. 7)',      _lev_flag(gross_lev,  fund_id, 'gross')),
        ('Commitment leverage (EU231/2013 Art. 8)', _lev_flag(commit_lev, fund_id, 'commitment')),
        ('', ''),
        ('LIQUIDITY HEADLINE', ''),
        ('% NAV liquidatable within 1 day', f'{pct_1d:.1f}%'),
    ]
    return pd.DataFrame(rows, columns=['field', 'value'])


def _build_leverage_detail(gross_lev: float, commit_lev: float, nav: float,
                           fund_id: str, breakdown: pd.DataFrame) -> pd.DataFrame:
    lims = _LEV_LIMITS[fund_id]

    rows = [('GROSS METHOD — by source (EU231/2013 Art. 7)', '', '')]
    for _, r in breakdown.iterrows():
        rows.append((r['source'], _eur(r['gross_eur']), f"{r['pct_nav']:.2f}% NAV"))
    rows += [
        ('Total gross exposure',        _eur(gross_lev * nav), f"{gross_lev:.2f}x NAV"),
        ('Gross leverage limit (RMP)',  f"{lims['gross']:.1f}x NAV", ''),
        ('', '', ''),
        ('COMMITMENT METHOD (EU231/2013 Art. 8)', '', ''),
        ('Net nettable exposure',           _eur(commit_lev * nav), f"{commit_lev:.2f}x NAV"),
        ('Commitment leverage limit (RMP)', f"{lims['commitment']:.1f}x NAV", ''),
    ]
    return pd.DataFrame(rows, columns=['item', 'gross_eur', 'pct_nav'])


def _aggregate_liquidity_buckets(liq_pos: pd.DataFrame, nav: float) -> pd.DataFrame:
    df = (liq_pos.groupby('liquidity_bucket', observed=True)['market_value_eur']
          .sum()
          .reindex(LIQUIDITY_BUCKET_ORDER, fill_value=0.0)
          .reset_index()
          .rename(columns={'liquidity_bucket': 'bucket', 'market_value_eur': 'nav_eur'}))
    df['nav_pct']        = (df['nav_eur'] / nav * 100).round(2) if nav else 0.0
    df['cumulative_pct'] = df['nav_pct'].cumsum().round(2)
    return df


def _build_liquidity_terms(fund_id: str, nav: float) -> pd.DataFrame:
    red     = _REDEMPTION[fund_id]
    inv_df  = pd.DataFrame(
        _INVESTOR_WEIGHTS.get(fund_id, []),
        columns=['investor_id', 'investor_name', 'investor_type', 'weight'])
    inv_df['aum_eur'] = inv_df['weight'] * nav

    ic     = investor_concentration(inv_df[['investor_id', 'investor_name', 'aum_eur']], nav)
    top1   = ic['largest_investor_pct'] * 100
    top3   = ic['top3_pct'] * 100
    c_flag = '⚠ Yes' if ic['concentration_flag'] else '✓ No'
    hc     = '⚠ Yes' if ic['high_concentration']  else '✓ No'

    rows = [
        ('Redemption frequency', red['frequency']),
        ('Notice period',        _fmt_red(red, 'notice_days')),
        ('Redemption gate',      _fmt_red(red, 'gate_pct')),
        ('Investor lock-up',     _fmt_red(red, 'lockup_months')),
        ('', ''),
        ('INVESTOR CONCENTRATION (ESMA thresholds)', ''),
        ('Largest single investor (% NAV)',  f'{top1:.1f}%'),
        ('Top 3 investors combined (% NAV)', f'{top3:.1f}%'),
        ('Single investor > 20% NAV flag',   c_flag),
        ('Top 3 > 50% NAV flag',             hc),
    ]
    return pd.DataFrame(rows, columns=['field', 'value'])


def _build_aifmd_ii(sub_limit: float, sub_drawn: float,
                    undrawn: float) -> pd.DataFrame:
    rows = [
        ('AIFMD II — EXPANDED DISCLOSURES (Directive 2024/927/EU)', ''),
        ('', ''),
        ('LIQUIDITY MANAGEMENT TOOLS (LMTs)', ''),
        ('Suspension of redemptions', 'Available — LP approval + board resolution required'),
        ('Side pockets',              'Available for impaired/illiquid positions — board decision'),
        ('Capital call facility',     f"{_eur(sub_limit)} limit / {_eur(sub_drawn)} drawn"),
        ('', ''),
        ('DELEGATION ARRANGEMENTS', ''),
        ('Portfolio management delegation', 'None — managed in-house by AIFM'),
        ('Risk management delegation',      'None — risk function internal to AIFM'),
        ('Valuation', 'Independent appraiser — KPMG / Duff & Phelps (AIFMD Art. 19)'),
        ('', ''),
        ('PRINCIPAL MARKETS', ''),
        ('Primary market',    'Direct (bilateral) investments'),
        ('Listed instruments', 'None — fully unlisted'),
        ('FX hedging',         'None at fund level'),
        ('', ''),
        ('UNFUNDED COMMITMENTS', ''),
        ('Total unfunded (contingent leverage)', _eur(undrawn)),
        ('Expected drawdown next 12 months',
            'Per capital call schedule — not disclosed in Annex IV'),
    ]
    return pd.DataFrame(rows, columns=['field', 'value'])


def _build_top5(rows_top5: list[dict], nav: float,
                value_col: str = 'nav_eur') -> pd.DataFrame:
    top5 = (pd.DataFrame(rows_top5)
            .sort_values(value_col, ascending=False).head(5))
    total = top5[value_col].sum()
    top5['nav_pct'] = (top5[value_col] / total * 100).round(2) if total else 0.0
    top5['rank']    = range(1, len(top5) + 1)
    top5[value_col] = top5[value_col].map('{:,.0f}'.format)

    return top5.reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════
# Fund-type builders
# ══════════════════════════════════════════════════════════════════════════

def _build_liquid(engine, fund_id: str, quarter: str) -> dict[str, pd.DataFrame]:
    risk_df  = get_risk_ready_df(engine, fund_id, _VALUATION_DATE)
    nav      = float(risk_df['market_value_eur'].sum())
    pnl      = query_nav_history(engine, fund_id)['pnl_pct'].dropna().values

    gross_lev, breakdown = _compute_gross_leverage(risk_df, nav)
    commit_lev           = _compute_commitment_leverage(risk_df, nav)

    liq = compute_liquidity_profile(risk_df, pct_adv=0.25)
    liq_df  = _aggregate_liquidity_buckets(liq['risk_df_liq'], nav)

    exposures = _build_exposures(risk_df, nav)

    # ── Breakdown section construction ────────────────────────────────────────
    cols = ['Category', 'NAV (EUR)', '% NAV']

    def _prep(df: pd.DataFrame, label: str) -> pd.DataFrame:
        d = df.copy().reset_index(drop=True)
        d.columns = cols
        d['NAV (EUR)'] = d['NAV (EUR)'].map('{:,.0f}'.format)
        d['% NAV'] = d['% NAV'].map('{:.2f}%'.format)
        header = pd.DataFrame([[label, '', '']], columns=cols)
        return pd.concat([header, d], ignore_index=True)

    df_top5 = exposures['top5_positions'][['name', 'market_value_eur', 'nav_pct']].copy()
    df_top5['name'] = df_top5['name'].str.replace(r'\(.*?\)', '', regex=True)
    df_top5.columns = cols

    sep = pd.DataFrame([['', '', '']], columns=cols)

    pieces = [
        _prep(exposures['asset_class_breakdown'], 'Asset Class'),
        sep,
        _prep(exposures['geography_breakdown'], 'Geography'),
        sep,
        _prep(exposures['currency_breakdown'], 'Currency'),
        sep,
        _prep(df_top5, 'Top 5 positions'),
    ]
    breakdown_df = pd.concat(pieces, ignore_index=True)

    return {
        'identification'       : _build_identification(engine, fund_id, quarter),
        'asset_class_breakdown': exposures['asset_class_breakdown'],
        'geography_breakdown'  : exposures['geography_breakdown'],
        'currency_breakdown'   : exposures['currency_breakdown'],
        'top5_positions'       : exposures['top5_positions'],
        'risk_measures'        : _build_risk_measures(pnl, nav, fund_id,
                                    gross_lev, commit_lev, liq_df),
        'leverage_detail'      : _build_leverage_detail(gross_lev, commit_lev,
                                    nav, fund_id, breakdown),
        'liquidity_buckets'    : liq_df,
        'liquidity_terms'      : _build_liquidity_terms(fund_id, nav),
        'breakdown'            : breakdown_df,
        '_nav'                 : nav,
    }


def _build_pe(engine, fund_id: str, quarter: str) -> dict[str, pd.DataFrame]:
    with Session(engine) as s:
        pef     = s.query(PEFund).filter_by(fund_id=fund_id).first()
        invs    = (s.query(PEFundInvestment, PEPortfolioCompany)
                   .join(PEPortfolioCompany,
                         PEFundInvestment.company_id == PEPortfolioCompany.company_id)
                   .filter(PEFundInvestment.fund_id == fund_id).all())
        nav_row = (s.query(PENavHistory)
                   .filter_by(fund_id=fund_id, company_id=None)
                   .order_by(PENavHistory.nav_date.desc()).first())
        cm      = (s.query(PEFundCashManagement)
                   .filter_by(fund_id=fund_id)
                   .order_by(PEFundCashManagement.cash_management_date.desc()).first())
        cfs     = s.query(PECashFlow).filter_by(fund_id=fund_id).all()

    nav        = float(nav_row.nav_eur) if nav_row else 0.0
    target_eur = float(pef.target_size_eur) if pef else 0.0
    drawn_eur  = sum(abs(c.amount_eur) for c in cfs
                     if c.amount_eur < 0 and c.flow_type == 'capital_call')
    undrawn    = max(0.0, target_eur - drawn_eur)
    sub_drawn  = float(cm.sub_line_drawn) if cm and cm.sub_line_drawn else 0.0
    sub_limit  = float(cm.sub_line_limit) if cm and cm.sub_line_limit else 0.0
    gross_lev  = (nav + sub_drawn) / nav if nav else 1.0
    commit_lev = gross_lev

    mult           = pe_multiples(engine, fund_id, as_of_date=quarter)
    irr_res        = fund_irr(engine, fund_id, as_of_date=quarter)
    life_remaining = _fund_life_remaining(pef)
    lims           = _LEV_LIMITS[fund_id]
    g_flag         = '⚠ BREACH' if gross_lev > lims['gross'] else '✓ within limit'

    rows_sector  = [{'sector':  co.sector  or 'Other',   'nav_eur': float(inv.cost_basis_eur)} for inv, co in invs]
    rows_country = [{'country': co.country or 'Unknown', 'nav_eur': float(inv.cost_basis_eur)} for inv, co in invs]
    rows_stage   = [{'stage':   co.investment_stage or 'Unknown', 'nav_eur': float(inv.cost_basis_eur)} for inv, co in invs]
    rows_top5    = [{'rank': 0, 'name': co.company_name, 'sector': co.sector or '',
                     'country': co.country or '', 'nav_eur': float(inv.cost_basis_eur)}
                    for inv, co in invs]

    perf_rows = (
        _fund_header_rows(pef, life_remaining,
                          extra=[('Strategy', pef.strategy if pef else '—'), ('', '')])
        + _capital_rows(target_eur, drawn_eur, nav, undrawn)
        + [('', '')]
        + _return_rows(mult, irr_res.get('net_irr', 0.0),
               gross_irr=irr_res.get('gross_irr', 0.0),
               moic_label='TVPI')
        + [('Paid-in capital', _eur(mult.get('paid_in', 0))),
           ('Distributions',   _eur(mult.get('distributions', 0)))]
    )

    lev_rows = [
        ('FUND-LEVEL LEVERAGE (EU231/2013 Art. 7)', '', ''),
        ('NAV',                                _eur(nav), ''),
        ('Subscription credit facility drawn', _eur(sub_drawn),
            f'{sub_drawn/nav*100:.1f}% NAV' if nav else '—'),
        ('Subscription credit facility limit', _eur(sub_limit), ''),
        ('Gross method leverage (fund level)', f'{gross_lev:.2f}x NAV',
            f'{g_flag}   limit: {lims["gross"]:.1f}x'),
    ] + _project_debt_note()

    return {
        'identification'     : _build_identification(engine, fund_id, quarter),
        'sector_exposure'    : _pct_df(rows_sector,  'sector').rename(columns={'nav_pct': 'cost_pct'}),
        'country_exposure'   : _pct_df(rows_country, 'country').rename(columns={'nav_pct': 'cost_pct'}),
        'stage_exposure'     : _pct_df(rows_stage,   'stage').rename(columns={'nav_pct': 'cost_pct'}),        'top5_positions'     : _build_top5(rows_top5, nav),
        'leverage_detail'    : pd.DataFrame(lev_rows, columns=['item', 'value', 'note']),
        'performance'        : pd.DataFrame(perf_rows, columns=['field', 'value']),
        'aifmd_ii_disclosure': _build_aifmd_ii(sub_limit, sub_drawn, undrawn),
        '_nav'               : nav,
    }


def _build_infra(engine, fund_id: str, quarter: str) -> dict[str, pd.DataFrame]:
    with Session(engine) as s:
        ifund   = s.query(InfraFund).filter_by(fund_id=fund_id).first()
        nav_row = (s.query(InfraNavHistory)
                   .filter_by(fund_id=fund_id, asset_id=None)
                   .order_by(InfraNavHistory.nav_date.desc()).first())
        invs    = (s.query(InfraFundInvestment, InfraAsset)
                   .join(InfraAsset,
                         InfraFundInvestment.asset_id == InfraAsset.asset_id)
                   .filter(InfraFundInvestment.fund_id == fund_id).all())
        proj_debt = (s.query(InfraDebt)
                     .filter(InfraDebt.asset_id.in_(
                         [inv.asset_id for inv, _ in invs]))
                     .all())

    nav           = float(nav_row.nav_eur)    if nav_row else 0.0
    committed_eur = float(ifund.committed_eur) if ifund and ifund.committed_eur else 0.0
    drawn_eur     = float(ifund.drawn_eur)     if ifund and ifund.drawn_eur     else 0.0
    total_debt    = sum(d.outstanding_eur or 0.0 for d in proj_debt)

    asset_bkd  = asset_nav_breakdown(engine, fund_id, quarter)
    sector_bkd = concentration_by_sector(engine, fund_id, quarter)
    dur_df     = duration_profile(engine, fund_id)
    infl       = inflation_sensitivity(engine, fund_id)
    mult       = infra_multiples(engine, fund_id)
    irr        = infra_irr(engine, fund_id)


    def _asset_nav(asset_id: str) -> float:
        if 'asset_id' not in asset_bkd.columns:
            return 0.0
        rows = asset_bkd.loc[asset_bkd['asset_id'] == asset_id, 'nav_eur']
        return float(rows.iloc[0]) if not rows.empty else 0.0

    rows_country = [{'country': a.country or 'Unknown', 'nav_eur': _asset_nav(a.asset_id)} for _, a in invs]
    rows_top5    = [{'rank': 0, 'name': a.asset_name, 'sector': a.sector or '',
                     'country': a.country or '', 'nav_eur': _asset_nav(a.asset_id)}
                    for _, a in invs]

    life_remaining = _fund_life_remaining(ifund)
    wad            = dur_df.attrs.get('weighted_avg_remaining_years')
    lims           = _LEV_LIMITS[fund_id]

    lev_rows = [
        ('FUND-LEVEL LEVERAGE (EU231/2013 Art. 7)', '', ''),
        ('Fund NAV',                          _eur(nav),           ''),
        ('Committed capital',                 _eur(committed_eur), ''),
        ('Drawn capital',                     _eur(drawn_eur),     ''),
        ('Fund-level debt',                   'None — no subscription credit line', ''),
        ('Gross method leverage (fund level)', '1.00x NAV',
            f'✓ within limit ({lims["gross"]:.1f}x)'),
        ('Commitment method leverage',         '1.00x NAV',
            f'✓ within limit ({lims["commitment"]:.1f}x)'),
        ('', '', ''),
        ('PROJECT-LEVEL DEBT (off-balance, disclosed separately)', '', ''),
        ('Total project-level debt outstanding', _eur(total_debt),
            'Ring-fenced at SPV level — excluded from AIFMD leverage'),
        ('Number of debt-bearing assets',
            str(len([d for d in proj_debt if d.outstanding_eur])),
            'Project finance structure'),
        ('AIFMD treatment',
            'Project debt excluded from Art. 7 gross leverage',
            'EU231/2013 Art. 7 — project finance carve-out'),
    ]

    perf_rows = (
        _fund_header_rows(ifund, life_remaining,
                          extra=[('Classification',
                                  ifund.aifmd_classification if ifund else '—'),
                                 ('', '')])
        + _capital_rows(committed_eur, drawn_eur, nav)
        + [('', '')]
        + _return_rows(mult, irr)
        + [
            ('', ''),
            ('INFRASTRUCTURE CHARACTERISTICS', ''),
            ('Number of assets',                    str(len(invs))),
            ('Weighted avg concession life (years)', f"{wad:.1f}" if wad else '—'),
            ('Weighted avg inflation linkage',
                f"{infl.get('weighted_avg_linkage', 0)*100:.1f}%"),
            ('% fully inflation-linked',
                f"{infl.get('pct_fully_linked', 0):.1f}%"),
            ('% partially inflation-linked',
                f"{infl.get('pct_partially_linked', 0):.1f}%"),
        ]
    )
    
    asset_bkd['nav_eur']  = asset_bkd['nav_eur'].map('{:,.0f}'.format)
    sector_bkd['nav_eur'] = sector_bkd['nav_eur'].map('{:,.0f}'.format)

    return {
        'identification'   : _build_identification(engine, fund_id, quarter),
        'asset_breakdown'  : asset_bkd,
        'sector_breakdown' : sector_bkd,
        'country_breakdown': _pct_df(rows_country, 'country'),
        'top5_positions'   : _build_top5(rows_top5, nav),
        'leverage_detail'  : pd.DataFrame(lev_rows, columns=['item', 'value', 'note']),
        'performance'      : pd.DataFrame(perf_rows, columns=['field', 'value']),
        '_nav'             : nav,
    }


# ══════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════

_BUILDERS = {
    'hf'   : _build_liquid,
    'pd'   : _build_liquid,
    're'   : _build_liquid,
    'pe'   : _build_pe,
    'infra': _build_infra,
}


def build_annex_iv(engine, fund_id: str,
                   quarter: str = '2026-03-31') -> dict[str, pd.DataFrame]:
    ftype = _FUND_TYPE.get(fund_id)
    if ftype is None:
        raise ValueError(f'Unknown fund_id: {fund_id}')
    return _BUILDERS[ftype](engine, fund_id, quarter)


# ══════════════════════════════════════════════════════════════════════════
# Excel helpers
# ══════════════════════════════════════════════════════════════════════════

_BG_HEADER  = '1a1f2e'
_BG_SECTION = '0f1729'
_BG_ALT     = '1d2235'
_BG_TOTAL   = '141929'
_FG_HEADER  = 'f9fafb'
_FG_ACCENT  = '1a9ed4'
_FG_WARN    = 'ef4444'
_FG_OK      = '22c55e'
_FG_MUTED   = '9ca3af'
_FG_DIM     = '6b7280'
_SIDE       = Side(style='thin', color='374151')
_BORDER     = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)


def _hfill(c: str) -> PatternFill:
    return PatternFill('solid', fgColor=c)

def _font(bold: bool = False, color: str = _FG_HEADER, size: int = 10) -> Font:
    return Font(name='Calibri', bold=bold, color=color, size=size)

def _align(h: str = 'left', v: str = 'center') -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=False)


def _write_header(ws, row: int, col: int, value: str, width: int = 1,
                  bold: bool = True, bg: str = _BG_HEADER,
                  fg: str = _FG_HEADER) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _hfill(bg)
    cell.font = _font(bold=bold, color=fg)
    cell.border = _BORDER
    cell.alignment = _align('center')
    if width > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + width - 1)


def _write_cell(ws, row: int, col: int, value: Any, bold: bool = False,
                bg: str = _BG_SECTION, fg: str = _FG_HEADER,
                align: str = 'left', num_fmt: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _hfill(bg)
    cell.font = _font(bold=bold, color=fg)
    cell.border = _BORDER
    cell.alignment = _align(align)
    if num_fmt:
        cell.number_format = num_fmt


def _write_field_value_df(ws, row: int, df: pd.DataFrame,
                          ncols: int = 2) -> int:
    for idx, r in df.iterrows():
        f, v = str(r.iloc[0]), str(r.iloc[1]) if len(r) > 1 else ''
        if f == '' and v == '':
            row += 1
            continue
        is_header = (v == '' and f == f.upper() and f != '')
        bg = _BG_HEADER if is_header else (_BG_ALT if idx % 2 == 0 else _BG_SECTION)
        fg = _FG_ACCENT if is_header else _FG_HEADER
        _write_cell(ws, row, 1, f, bold=is_header, bg=bg, fg=fg)
        if ncols >= 2:
            _write_cell(ws, row, 2, v, bg=bg, fg=_FG_HEADER)
        row += 1
    return row + 1


def _write_df_table(ws, row: int, df: pd.DataFrame,
                    title: str = '', alt: bool = True) -> int:
    if title:
        _write_header(ws, row, 1, title, width=len(df.columns), bg=_BG_HEADER)
        row += 1
    for ci, col in enumerate(df.columns, start=1):
        _write_header(ws, row, ci, str(col), width=1, bg='262d40')
    row += 1
    for idx, r in df.iterrows():
        bg = _BG_ALT if (alt and idx % 2 == 0) else _BG_SECTION
        for ci, val in enumerate(r, start=1):
            _write_cell(ws, row, ci, val, bg=bg)
        row += 1
    return row + 1


_SHEET_NAMES = {
    'AIFM_HedgeFund'  : 'HedgeFund',
    'AIFM_PrivateDebt': 'PrivateDebt',
    'AIFM_RealEstate' : 'RealEstate',
    'AIFM_PE_Buyout'  : 'PEBuyout',
    'AIFM_Infra_Core' : 'InfraCore',
}

_EXPORT_FUNDS = list(_SHEET_NAMES.keys())


def _write_fund_sheet(engine: sa.Engine, wb: Workbook, fund_id: str,
                      rpt: dict[str, pd.DataFrame], quarter: str) -> None:
    ws = wb.create_sheet(title=_SHEET_NAMES[fund_id])
    ws.sheet_view.showGridLines = False
    for col, width in zip('ABCDE', [48, 38, 32, 16, 14]):
        ws.column_dimensions[col].width = width

    nav  = rpt.get('_nav', 0.0)
    meta = _get_fund_metadata(engine, fund_id)

    _write_header(ws, 1, 1,
                  f"Annex IV Regulatory Report — {meta.get('fund_name', fund_id)}",
                  width=5, bg=_BG_HEADER)
    _write_header(ws, 2, 1,
                  f"Fund: {fund_id}   |   NAV: EUR {nav/1e6:,.1f}M   "
                  f"|   Reporting period: Q1 2026 ({quarter})   "
                  f"|   Generated: {VALUATION_DATE}",
                  width=5, bg=_BG_SECTION, bold=False, fg=_FG_MUTED)
    _write_header(ws, 3, 1,
                  'Regulatory basis: AIFMD Art. 110 / EU231/2013 Annex IV / '
                  'ESMA technical guidance v1.7 (July 2024)',
                  width=5, bg=_BG_SECTION, bold=False, fg=_FG_DIM)

    row   = 5
    ftype = _FUND_TYPE[fund_id]
    row   = _write_field_value_df(ws, row, rpt['identification'])

    if ftype in ('hf', 'pd', 're'):
        for key, title in [
            ('asset_class_breakdown', 'Section 2.1 — Asset class breakdown'),
            ('geography_breakdown',   'Section 2.2 — Geographic breakdown (top 10)'),
            ('currency_breakdown',    'Section 2.3 — Currency breakdown'),
            ('top5_positions',        'Section 2.4 — Top 5 positions by absolute exposure'),
        ]:
            row = _write_df_table(ws, row, rpt[key], title)
        row = _write_field_value_df(ws, row, rpt['risk_measures'])
        row = _write_df_table(ws, row, rpt['leverage_detail'], 'Section 4 — Leverage detail')
        row = _write_df_table(ws, row, rpt['liquidity_buckets'], 'Section 5.1 — Liquidity profile (ESMA buckets)')
        row = _write_field_value_df(ws, row, rpt['liquidity_terms'])

    elif ftype == 'pe':
        for key, title in [
            ('sector_exposure',  'Section 2.1 — Portfolio exposure by sector (cost basis)'),
            ('country_exposure', 'Section 2.2 — Portfolio exposure by country'),
            ('stage_exposure',   'Section 2.3 — Portfolio exposure by investment stage'),
            ('top5_positions',   'Section 2.4 — Top 5 investments'),
            ('leverage_detail',  'Section 3 — Leverage (fund-level)'),
        ]:
            row = _write_df_table(ws, row, rpt[key], title)
        row = _write_field_value_df(ws, row, rpt['performance'])
        row = _write_field_value_df(ws, row, rpt['aifmd_ii_disclosure'])

    elif ftype == 'infra':
        for key, title in [
            ('asset_breakdown',   'Section 2.1 — Asset NAV breakdown'),
            ('sector_breakdown',  'Section 2.2 — Sector concentration'),
            ('country_breakdown', 'Section 2.3 — Geographic breakdown'),
            ('top5_positions',    'Section 2.4 — Top 5 assets by NAV'),
            ('leverage_detail',   'Section 3 — Leverage'),
        ]:
            row = _write_df_table(ws, row, rpt[key], title)
        row = _write_field_value_df(ws, row, rpt['performance'])

    ws.freeze_panes = 'A5'


def _write_summary_sheet(engine: sa.Engine, wb: Workbook,
                         reports: dict[str, dict],
                         quarter: str) -> None:
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False
    for col, width in zip('ABCDEF', [40, 20, 20, 20, 20, 20]):
        ws.column_dimensions[col].width = width

    funds = list(reports.keys())
    nf    = len(funds)

    _write_header(ws, 1, 1, 'Annex IV Regulatory Report — Cross-Fund Summary',
                  width=1 + nf, bg=_BG_HEADER)
    _write_header(ws, 2, 1,
                  f'Reporting period: Q1 2026 ({quarter})   |   '
                  f'Generated: {VALUATION_DATE}   |   '
                  'Regulatory basis: AIFMD Art. 110 / EU231/2013 Annex IV',
                  width=1 + nf, bg=_BG_SECTION, bold=False, fg=_FG_MUTED)

    _write_header(ws, 4, 1, 'Key metric', bg=_BG_HEADER)
    for i, fid in enumerate(funds, start=2):
        _write_header(ws, 4, i,
                      _get_fund_metadata(engine, fid).get('fund_name', fid),
                      bg=_BG_HEADER)

    summary_labels = [
        'NAV (EUR M)', 'Strategy', 'Geography',
        'Gross leverage', 'Commitment leverage',
        'VaR 1-day 99% (%)', 'VaR 20-day 99% (%)',
        'Redemption frequency', 'Fund life remaining (years)',
    ]

    for si, label in enumerate(summary_labels):
        row = 5 + si
        bg  = _BG_ALT if si % 2 == 0 else _BG_SECTION
        _write_cell(ws, row, 1, label, bold=True, bg=bg, align='left')

        for i, fid in enumerate(funds, start=2):
            rpt   = reports[fid]
            ftype = _FUND_TYPE[fid]
            nav   = rpt.get('_nav', 0.0)
            val   = '—'

            if label == 'NAV (EUR M)':
                val = f"EUR {nav/1e6:,.1f}M"
            elif label == 'Strategy':
                val = _STRATEGY.get(fid, '—')
            elif label == 'Geography':
                val = _GEO.get(fid, '—')
            elif label == 'Gross leverage':
                lev_df = rpt.get('leverage_detail', pd.DataFrame())
                if 'item' in lev_df.columns:
                    r = lev_df[lev_df['item'].str.contains('Gross method leverage', na=False)]
                    val = str(r['value'].iloc[0]) if not r.empty else '—'
            elif label == 'Commitment leverage':
                lev_df = rpt.get('leverage_detail', pd.DataFrame())
                if 'item' in lev_df.columns:
                    key = 'Net nettable' if ftype in ('hf', 'pd', 're') else 'Commitment method'
                    r   = lev_df[lev_df['item'].str.contains(key, na=False)]
                    col = 'pct_nav' if ftype in ('hf', 'pd', 're') else 'value'
                    val = str(r[col].iloc[0]) if not r.empty else '—'
            elif label in ('VaR 1-day 99% (%)', 'VaR 20-day 99% (%)'):
                if ftype in ('hf', 'pd', 're'):
                    rm  = rpt.get('risk_measures', pd.DataFrame())
                    key = 'VaR 1-day (99%)' if '1-day' in label else 'VaR 20-day (99%)'
                    r   = rm[rm['field'] == key] if 'field' in rm.columns else pd.DataFrame()
                    val = str(r['value'].iloc[0]).split()[0] if not r.empty else '—'
                else:
                    val = 'N/A — illiquid'
            elif label == 'Redemption frequency':
                val = _REDEMPTION.get(fid, {}).get('frequency', '—')
            elif label == 'Fund life remaining (years)':
                if ftype in ('pe', 'infra'):
                    perf = rpt.get('performance', pd.DataFrame())
                    if 'field' in perf.columns:
                        r = perf[perf['field'] == 'Fund life remaining (years)']
                        val = str(r['value'].iloc[0]) if not r.empty else '—'
                else:
                    val = 'Open-ended'

            _write_cell(ws, row, i, val, bg=bg, align='right')

    ws.freeze_panes = 'A5'


# ══════════════════════════════════════════════════════════════════════════
# Excel export
# ══════════════════════════════════════════════════════════════════════════

def export_annex_iv_excel(
    engine=None,
    quarter: str = '2026-03-31',
    output_dir: str = 'data',
    fund_ids: list[str] | None = None,
) -> str:
    if engine is None:
        engine = get_engine()
    if fund_ids is None:
        fund_ids = _EXPORT_FUNDS

    # Resolve output directory and compute paths
    from pathlib import Path
    from datetime import datetime
    from fund_risk_workflow.data.paths import annex_iv_file, annex_iv_dir

    project_root = Path(__file__).parent.parent.parent.parent
    out_path_obj = Path(output_dir)

    # If output_dir is relative, resolve it from the project root
    if not out_path_obj.is_absolute():
        resolved_path = (project_root / output_dir).resolve()
    else:
        resolved_path = out_path_obj.resolve()

    # Convert quarter from YYYY-MM-DD to YYYYQN format (e.g., 2026-03-31 -> 2026Q1)
    quarter_date = datetime.strptime(quarter, '%Y-%m-%d')
    quarter_num = (quarter_date.month - 1) // 3 + 1
    quarter_formatted = f'{quarter_date.year}Q{quarter_num}'

    # Determine fund label for filename
    if len(fund_ids) == len(_EXPORT_FUNDS):
        # All funds case
        fund_label = None  # Will use 'all_funds.xlsx'
    elif len(fund_ids) == 1:
        # Single fund case
        fund_label = fund_ids[0]
    else:
        # Multiple specific funds case
        fund_label = None  # Will use 'all_funds.xlsx'

    # Use path helper to construct file path
    out_path = str(annex_iv_file(str(resolved_path), quarter_formatted, fund_label))
    # Ensure directory exists
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)

    # Only print progress for all-funds export
    is_all_funds = len(fund_ids) == len(_EXPORT_FUNDS)
    if is_all_funds:
        print(f'Annex IV export')
        print(f'Reporting period: {quarter_formatted}')

    reports: dict[str, dict] = {}
    for fid in fund_ids:
        if is_all_funds:
            print(f'  Building {fid}...', end=' ', flush=True)
        rpt = build_annex_iv(engine, fid, quarter)
        reports[fid] = rpt
        if is_all_funds:
            print(f'NAV EUR {rpt.get("_nav", 0)/1e6:,.1f}M')

    wb = Workbook()
    _write_summary_sheet(engine, wb, reports, quarter)
    for fid in fund_ids:
        _write_fund_sheet(engine, wb, fid, reports[fid], quarter)

    wb.save(out_path)
    # Return and display relative path from project root
    relative_path = Path(out_path).relative_to(project_root)
    print(f'\nReport saved: {relative_path}')
    return str(relative_path)


if __name__ == '__main__':
    export_annex_iv_excel()