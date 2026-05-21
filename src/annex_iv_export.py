"""
src/annex_vi_export.py
======================
Annex VI stress test Excel export — CSSF submission format.

Runs all regulatory stress scenarios for the three liquid AIFM funds
(Hedge Fund, Private Debt, Real Estate) and writes a multi-sheet workbook
suitable for direct CSSF submission.

Usage
-----
    from src.annex_iv_export import generate_annex_iv_report
    generate_annex_iv_report(engine, valuation_date='2026-05-13')

    # or from the command line:
    python -m src.annex_iv_export

Output
------
    data/annex_iv_report_<date>.xlsx

    Sheets
    ------
    Summary            Cross-fund scenario comparison (all funds side by side)
    AIFM_HedgeFund     Detailed HF stress results
    AIFM_PrivateDebt   Detailed PD stress results
    AIFM_RealEstate    Detailed RE stress results

Regulatory basis
----------------
    ESMA/2020/1498 (Annex VI) — stress testing guidelines for AIFMs
    AIFMD Article 15 / Annex IV — risk management and reporting requirements
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Dict, List, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from src.database import get_engine
from src.enrichment import get_risk_ready_df
from src.risk_utils import (
    HISTORICAL_SCENARIOS,
    stress_equity, stress_rates, stress_credit, stress_fx,
    stress_combined, stress_historical,
    stress_property, stress_rental,
    days_to_liquidate, liquidity_buckets, redemption_stress,
)

# ── colour palette (dark theme consistent with plot_style) ────────────────
_BG_HEADER  = '1a1f2e'   # dark navy
_BG_SECTION = '0f1729'   # darker
_BG_TOTAL   = '141929'
_FG_HEADER  = 'f9fafb'   # near-white
_FG_ACCENT  = '1a9ed4'   # cyan accent
_FG_WARN    = 'ef4444'   # red for breaches
_FG_OK      = '22c55e'   # green for pass
_BG_ALT     = '1d2235'   # alternating row

_SIDE  = Side(style='thin', color='374151')
_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)


def _hfill(hex_code: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_code)


def _font(bold: bool = False, color: str = _FG_HEADER, size: int = 10) -> Font:
    return Font(name='Calibri', bold=bold, color=color, size=size)


def _align(h: str = 'left', v: str = 'center') -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=False)


def _write_header(ws, row: int, col: int, value: str,
                  width: int = 1, bold: bool = True,
                  bg: str = _BG_HEADER, fg: str = _FG_HEADER) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = _hfill(bg)
    cell.font   = _font(bold=bold, color=fg)
    cell.border = _BORDER
    cell.alignment = _align('center')
    if width > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + width - 1
        )


def _write_cell(ws, row: int, col: int, value: Any,
                bold: bool = False,
                bg: str = _BG_SECTION,
                fg: str = _FG_HEADER,
                align: str = 'right',
                num_fmt: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = _hfill(bg)
    cell.font      = _font(bold=bold, color=fg)
    cell.border    = _BORDER
    cell.alignment = _align(align)
    if num_fmt:
        cell.number_format = num_fmt


# ══════════════════════════════════════════════════════════════════════════
# stress runner helpers
# ══════════════════════════════════════════════════════════════════════════

def _run_scenarios_hf(risk_df: pd.DataFrame, nav: float) -> List[dict]:
    rows = []

    def _add(label: str, pnl_eur: float, category: str) -> None:
        rows.append({
            'category': category,
            'scenario': label,
            'pnl_eur' : pnl_eur,
            'pnl_pct' : pnl_eur / nav,
        })

    r = stress_equity(risk_df, delta_equity=-0.30)
    _add('Equity −30%',        r['stressed_pnl_eur'], 'Market')
    r = stress_equity(risk_df, delta_equity=-0.20)
    _add('Equity −20%',        r['stressed_pnl_eur'], 'Market')
    r = stress_rates(risk_df, delta_y=0.02)
    _add('Rates +200bps',      r['stressed_pnl_eur'], 'Market')
    r = stress_credit(risk_df, delta_spread=0.015)
    _add('Credit +150bps',     r['stressed_pnl_eur'], 'Market')
    r = stress_fx(risk_df, fx_shocks={'USD': -0.15, 'GBP': -0.15})
    _add('FX −15% (USD/GBP)',  r['stressed_pnl_eur'], 'Market')
    r = stress_combined(risk_df)
    _add('Combined (ESMA)',    r['stressed_pnl_eur'], 'Market')

    for yr, sc in HISTORICAL_SCENARIOS.items():
        r = stress_historical(risk_df, yr)
        _add(f'Historical {yr} ({sc["name"].split("(")[0].strip()})',
             r['stressed_pnl_eur'], 'Historical')

    # counterparty stress: GS PB defaults, 12% exposure, 80% collateral
    cp_loss = nav * 0.12 * (1 - 0.80)
    _add('Counterparty (GS default, 80% coll.)', -cp_loss, 'Counterparty')

    # combined stress: equity −20% + 25% redemption
    risk_liq = days_to_liquidate(risk_df, pct_adv=0.25)
    risk_liq = liquidity_buckets(risk_liq)
    r_eq     = stress_equity(risk_df, delta_equity=-0.20)
    r_red    = redemption_stress(risk_liq, nav, redemption_pct=0.25)
    liq_st   = r_red['liquid_assets_eur'] * 0.80   # liquid assets −20%
    shortfall = max(0.0, nav * 0.25 - liq_st)
    combined_pnl = r_eq['stressed_pnl_eur'] - shortfall
    _add('Combined (Equity −20% + 25% Redemption)', combined_pnl, 'Combined')

    return rows


def _run_scenarios_pd(risk_df: pd.DataFrame, nav: float) -> List[dict]:
    rows = []

    def _add(label: str, pnl_eur: float, category: str) -> None:
        rows.append({'category': category, 'scenario': label,
                     'pnl_eur': pnl_eur, 'pnl_pct': pnl_eur / nav})

    r = stress_rates(risk_df, delta_y=0.02)
    _add('Rates +200bps',      r['stressed_pnl_eur'], 'Market')
    r = stress_credit(risk_df, delta_spread=0.015)
    _add('Credit +150bps',     r['stressed_pnl_eur'], 'Market')
    r = stress_combined(risk_df)
    _add('Combined (ESMA)',    r['stressed_pnl_eur'], 'Market')

    for yr, sc in HISTORICAL_SCENARIOS.items():
        r = stress_historical(risk_df, yr)
        _add(f'Historical {yr} ({sc["name"].split("(")[0].strip()})',
             r['stressed_pnl_eur'], 'Historical')

    # largest borrower default: find from positions, 40% recovery
    loans = risk_df[risk_df['market_value_eur'] > 0]
    if not loans.empty and 'issuer' in loans.columns:
        by_b = loans.groupby('issuer', dropna=True)['market_value_eur'].sum()
        largest_exp = by_b.max() if not by_b.empty else 0.0
        borrow_loss = largest_exp * 0.60  # LGD 60%
        _add('Counterparty (largest borrower, 40% recovery)',
             -borrow_loss, 'Counterparty')
    else:
        _add('Counterparty (largest borrower, 40% recovery)', 0.0, 'Counterparty')

    # combined: credit +150bps + 25% redemption
    risk_liq = days_to_liquidate(risk_df, pct_adv=0.25)
    risk_liq = liquidity_buckets(risk_liq)
    cr        = stress_credit(risk_df, delta_spread=0.015)
    r_red     = redemption_stress(risk_liq, nav, redemption_pct=0.25)
    liq_st    = r_red['liquid_assets_eur'] * 0.90   # 10% haircut
    shortfall = max(0.0, nav * 0.25 - liq_st)
    combined_pnl = cr['stressed_pnl_eur'] - shortfall
    _add('Combined (Credit +150bps + 25% Redemption)', combined_pnl, 'Combined')

    return rows


def _run_scenarios_re(risk_df: pd.DataFrame, nav: float) -> List[dict]:
    rows = []

    def _add(label: str, pnl_eur: float, category: str) -> None:
        rows.append({'category': category, 'scenario': label,
                     'pnl_eur': pnl_eur, 'pnl_pct': pnl_eur / nav})

    r = stress_property(risk_df, delta_value_by_type={
        'Office': -0.20, 'Logistics': -0.20, 'Retail': -0.25, 'Residential': -0.15
    })
    _add('Property −20% (uniform)',        r['stressed_pnl_eur'], 'Market')
    r = stress_rental(risk_df, delta_vacancy=0.10, delta_yield=-0.005)
    _add('Rental stress (+10% vacancy)',   r['stressed_pnl_eur'], 'Market')
    r = stress_rates(risk_df, delta_y=0.02)
    _add('Rates +200bps',                  r['stressed_pnl_eur'], 'Market')
    r = stress_combined(risk_df)
    _add('Combined (ESMA)',                r['stressed_pnl_eur'], 'Market')

    for yr, sc in HISTORICAL_SCENARIOS.items():
        r = stress_historical(risk_df, yr)
        _add(f'Historical {yr} ({sc["name"].split("(")[0].strip()})',
             r['stressed_pnl_eur'], 'Historical')

    # largest tenant default: 1-year income, capitalised at 5% yield
    tenant_income_loss = 3_200_000          # Carrefour SA (largest simulated tenant)
    tenant_nav_impact  = -(tenant_income_loss / 0.05)  # capitalised
    _add('Counterparty (largest tenant default, capitalised)',
         tenant_nav_impact, 'Counterparty')

    # combined: property −20% + 25% redemption
    risk_liq = days_to_liquidate(risk_df, pct_adv=0.25)
    risk_liq = liquidity_buckets(risk_liq)
    prop_loss = stress_property(risk_df, delta_value_by_type={
        'Office': -0.20, 'Logistics': -0.20, 'Retail': -0.25, 'Residential': -0.15
    })['stressed_pnl_eur']
    r_red     = redemption_stress(risk_liq, nav, redemption_pct=0.25)
    liq_st    = r_red['liquid_assets_eur']    # cash unaffected by property shock
    shortfall = max(0.0, nav * 0.25 - liq_st)
    combined_pnl = prop_loss - shortfall
    _add('Combined (Property −20% + 25% Redemption)', combined_pnl, 'Combined')

    return rows


# ══════════════════════════════════════════════════════════════════════════
# Excel writer helpers
# ══════════════════════════════════════════════════════════════════════════

_CATS_ORDER = ['Market', 'Historical', 'Counterparty', 'Combined']

_COL_WIDTHS = {
    1: 28,   # Category
    2: 48,   # Scenario
    3: 18,   # ΔNAV (EUR M)
    4: 14,   # ΔNAV (% NAV)
    5: 20,   # Flag
}


def _write_fund_sheet(wb: Workbook, sheet_name: str,
                      fund_id: str, fund_label: str,
                      scenarios: List[dict], nav: float,
                      valuation_date: str) -> None:
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    for col, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title row
    _write_header(ws, 1, 1, f'Annex VI Stress Test Report — {fund_label}',
                  width=5, bg=_BG_HEADER)
    _write_header(ws, 2, 1,
                  f'Fund: {fund_id}  |  NAV: EUR {nav/1e6:,.1f}M  '
                  f'|  Valuation date: {valuation_date}  '
                  f'|  Generated: {datetime.today().strftime("%Y-%m-%d")}',
                  width=5, bg=_BG_SECTION, bold=False, fg='9ca3af')
    _write_header(ws, 3, 1,
                  'Regulatory basis: ESMA/2020/1498 Annex VI — quarterly stress testing',
                  width=5, bg=_BG_SECTION, bold=False, fg='6b7280')

    # Column headers
    r = 5
    for col, label in [(1, 'Category'), (2, 'Scenario'),
                       (3, 'ΔNAV (EUR M)'), (4, 'ΔNAV (% NAV)'),
                       (5, 'Regulatory flag')]:
        _write_header(ws, r, col, label, bg=_BG_HEADER)

    r += 1
    worst_pnl = float('inf')

    for cat in _CATS_ORDER:
        cat_rows = [s for s in scenarios if s['category'] == cat]
        if not cat_rows:
            continue

        for idx, s in enumerate(cat_rows):
            row_bg = _BG_ALT if idx % 2 == 0 else _BG_SECTION
            pnl_m  = s['pnl_eur'] / 1e6
            pnl_p  = s['pnl_pct'] * 100

            flag_txt  = '—'
            flag_color = _FG_HEADER
            if s['pnl_pct'] < -0.20:
                flag_txt, flag_color = '⚠ Severe (>20% NAV)', _FG_WARN
            elif s['pnl_pct'] < -0.10:
                flag_txt, flag_color = '⚠ Significant (>10% NAV)', 'f97316'
            elif s['pnl_pct'] < 0:
                flag_txt, flag_color = '✓ Manageable', _FG_OK
            else:
                flag_txt, flag_color = '✓ No loss', _FG_OK

            _write_cell(ws, r, 1, cat,          align='left',  bg=row_bg)
            _write_cell(ws, r, 2, s['scenario'], align='left',  bg=row_bg)
            _write_cell(ws, r, 3, round(pnl_m, 2),
                        bg=row_bg, align='right',
                        fg=_FG_WARN if pnl_m < 0 else _FG_OK)
            _write_cell(ws, r, 4, f'{pnl_p:.1f}%',
                        bg=row_bg, align='right',
                        fg=_FG_WARN if pnl_p < 0 else _FG_OK)
            _write_cell(ws, r, 5, flag_txt, bg=row_bg,
                        align='left', fg=flag_color)

            if s['pnl_eur'] < worst_pnl:
                worst_pnl = s['pnl_eur']
            r += 1

        # category divider
        r += 1

    # Worst-case summary
    r += 1
    _write_header(ws, r, 1, 'Worst-case scenario across all categories',
                  width=5, bg=_BG_TOTAL)
    r += 1
    _write_cell(ws, r, 1, 'WORST CASE', bold=True, align='left', bg=_BG_TOTAL)
    _write_cell(ws, r, 2, 'See scenario marked above', align='left', bg=_BG_TOTAL)
    _write_cell(ws, r, 3, round(worst_pnl / 1e6, 2), bg=_BG_TOTAL,
                align='right', fg=_FG_WARN)
    _write_cell(ws, r, 4, f'{worst_pnl/nav*100:.1f}%', bg=_BG_TOTAL,
                align='right', fg=_FG_WARN)
    _write_cell(ws, r, 5, '— Review risk management policy', bg=_BG_TOTAL,
                align='left', fg='f97316')

    # Freeze header rows
    ws.freeze_panes = 'A6'


def _write_summary_sheet(wb: Workbook,
                         fund_data: Dict[str, Dict],
                         valuation_date: str) -> None:
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    funds   = list(fund_data.keys())
    n_funds = len(funds)

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 48
    for i in range(n_funds):
        col_idx = 3 + i * 2
        ws.column_dimensions[get_column_letter(col_idx)].width     = 16
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 13

    # Title
    _write_header(ws, 1, 1,
                  'Annex VI Stress Test Report — Cross-Fund Summary',
                  width=2 + n_funds * 2, bg=_BG_HEADER)
    _write_header(ws, 2, 1,
                  f'Valuation date: {valuation_date}  |  '
                  f'Generated: {datetime.today().strftime("%Y-%m-%d")}  |  '
                  'Regulatory basis: ESMA/2020/1498 Annex VI',
                  width=2 + n_funds * 2, bg=_BG_SECTION, bold=False, fg='9ca3af')

    # Column headers
    r = 4
    _write_header(ws, r, 1, 'Category', bg=_BG_HEADER)
    _write_header(ws, r, 2, 'Scenario',  bg=_BG_HEADER)
    for i, fid in enumerate(funds):
        col = 3 + i * 2
        label = fund_data[fid]['label']
        nav_m = fund_data[fid]['nav'] / 1e6
        _write_header(ws, r, col,     f'{label}\nΔNAV (EUR M)', bg=_BG_HEADER)
        _write_header(ws, r, col + 1, f'{label}\nΔNAV (% NAV)', bg=_BG_HEADER)
        ws.row_dimensions[r].height = 30

    # collect all unique (category, scenario) pairs preserving order
    seen: list[tuple[str, str]] = []
    for cat in _CATS_ORDER:
        for fid in funds:
            for s in fund_data[fid]['scenarios']:
                if s['category'] == cat and (cat, s['scenario']) not in seen:
                    seen.append((cat, s['scenario']))

    r = 5
    for idx, (cat, scen) in enumerate(seen):
        row_bg = _BG_ALT if idx % 2 == 0 else _BG_SECTION
        _write_cell(ws, r, 1, cat,  align='left', bg=row_bg)
        _write_cell(ws, r, 2, scen, align='left', bg=row_bg)

        for i, fid in enumerate(funds):
            col  = 3 + i * 2
            nav  = fund_data[fid]['nav']
            match = next(
                (s for s in fund_data[fid]['scenarios']
                 if s['category'] == cat and s['scenario'] == scen),
                None
            )
            if match:
                pnl_m = round(match['pnl_eur'] / 1e6, 2)
                pnl_p = f"{match['pnl_pct']*100:.1f}%"
                fg_m = _FG_WARN if pnl_m < 0 else _FG_OK
                fg_p = _FG_WARN if match['pnl_pct'] < 0 else _FG_OK
                _write_cell(ws, r, col,     pnl_m, align='right', bg=row_bg, fg=fg_m)
                _write_cell(ws, r, col + 1, pnl_p, align='right', bg=row_bg, fg=fg_p)
            else:
                _write_cell(ws, r, col,     '—', align='center', bg=row_bg, fg='6b7280')
                _write_cell(ws, r, col + 1, '—', align='center', bg=row_bg, fg='6b7280')

        r += 1

    # Worst-case row per fund
    r += 1
    _write_header(ws, r, 1, 'Worst case', width=2, bg=_BG_TOTAL)
    for i, fid in enumerate(funds):
        col = 3 + i * 2
        nav = fund_data[fid]['nav']
        wc  = min(fund_data[fid]['scenarios'], key=lambda s: s['pnl_eur'])
        pnl_m = round(wc['pnl_eur'] / 1e6, 2)
        pnl_p = f"{wc['pnl_pct']*100:.1f}%"
        _write_cell(ws, r, col,     pnl_m, bold=True, bg=_BG_TOTAL, fg=_FG_WARN, align='right')
        _write_cell(ws, r, col + 1, pnl_p, bold=True, bg=_BG_TOTAL, fg=_FG_WARN, align='right')

    ws.freeze_panes = 'A5'


# ══════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════

_FUNDS = {
    'AIFM_HedgeFund':   {'label': 'HF',  'runner': _run_scenarios_hf},
    'AIFM_PrivateDebt': {'label': 'PD',  'runner': _run_scenarios_pd},
    'AIFM_RealEstate':  {'label': 'RE',  'runner': _run_scenarios_re},
}

_SHEET_NAMES = {
    'AIFM_HedgeFund':   'AIFM_HedgeFund',
    'AIFM_PrivateDebt': 'AIFM_PrivateDebt',
    'AIFM_RealEstate':  'AIFM_RealEstate',
}

_FUND_LABELS = {
    'AIFM_HedgeFund':   'AIFM Hedge Fund',
    'AIFM_PrivateDebt': 'AIFM Private Debt',
    'AIFM_RealEstate':  'AIFM Real Estate',
}


def generate_annex_iv_report(
    engine=None,
    valuation_date: str = '2026-05-13',
    output_dir: str = 'data',
) -> str:
    """
    Run all Annex VI stress scenarios for the three liquid AIFM funds
    and write a CSSF-ready Excel workbook.

    Parameters
    ----------
    engine : sqlalchemy Engine, optional
        Defaults to get_engine().
    valuation_date : str
        ISO date string. Must exist in the positions table.
    output_dir : str
        Directory to write the .xlsx file.

    Returns
    -------
    str
        Full path to the written workbook.
    """
    if engine is None:
        engine = get_engine()

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir,
                            f'annex_iv_report_{valuation_date}.xlsx')

    print(f'Annex VI export — valuation date {valuation_date}')

    fund_data: Dict[str, Dict] = {}
    for fund_id, cfg in _FUNDS.items():
        print(f'  Loading {fund_id}...', end=' ', flush=True)
        risk_df = get_risk_ready_df(engine, fund_id, valuation_date)
        nav     = float(risk_df['market_value_eur'].sum())
        rows    = cfg['runner'](risk_df, nav)
        fund_data[fund_id] = {
            'label'    : cfg['label'],
            'nav'      : nav,
            'scenarios': rows,
        }
        worst = min(rows, key=lambda r: r['pnl_eur'])
        print(f'NAV EUR {nav/1e6:,.1f}M  |  worst: {worst["pnl_pct"]*100:.1f}%')

    wb = Workbook()
    _write_summary_sheet(wb, fund_data, valuation_date)

    for fund_id, cfg in _FUNDS.items():
        fd = fund_data[fund_id]
        _write_fund_sheet(
            wb,
            sheet_name    = _SHEET_NAMES[fund_id],
            fund_id       = fund_id,
            fund_label    = _FUND_LABELS[fund_id],
            scenarios     = fd['scenarios'],
            nav           = fd['nav'],
            valuation_date= valuation_date,
        )

    wb.save(out_path)
    print(f'\nWritten: {out_path}')
    return out_path


if __name__ == '__main__':
    generate_annex_iv_report()
